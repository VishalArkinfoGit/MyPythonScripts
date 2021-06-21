from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime
import string

now = datetime.now()

my_path = "Documents\OneTree_AUS_20052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""

listOBJ = []
listError = []

url = 'https://www.onetree.org.au/locations/'

print(url)

try:

        res = requests.request("GET", url)

        if res.status_code == 200:
            soup = BeautifulSoup(res.content, "html.parser")

            try:
                output = str(soup).replace('\n','').replace('\t','').replace('  ','').strip()

                try:
                    index1 = output.index('var markers = [')
                    index2 = output.index('];', index1)
                    xz = output[index1 + len('var markers = ['):index2]

                    listLatLng = xz.split('],[')

                    for x in listLatLng:
                        obj = OBJ()

                        x= x.replace('[','').replace('\'','').replace('],','')

                        xy=x.split(',')

                        obj.Title = xy[0]
                        obj.Latitude = xy[1]
                        obj.Longitude = xy[2]

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.FullAddress) == str(
                                        listOBJ[i].FullAddress)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)

                except:
                    print("Nothing")


                try:
                    index1 = output.index('var infoWindowContent = [')
                    index2 = output.index('];', index1)
                    xz = output[index1 + len('var infoWindowContent = ['):index2]

                    listAddress = xz.split('],')

                    for x in listAddress:
                        x = x.replace('[', '').replace('\'', '').replace('],', '')

                        index1 = x.index('<h3>')
                        index2 = x.index('</h3>', index1)
                        name = x[index1 + len('<h3>'):index2]

                        index1 = x.index('<p>',index2)
                        index2 = x.index('</p>', index1)
                        address = x[index1 + len('<p>'):index2]

                        index = -1
                        try:
                            index = [x.Title for x in listOBJ].index(name)
                        except ValueError:
                            index = -1


                        if index > -1:
                            listOBJ[index].Address = address
                except:
                    print("Nothing")
            except:
                print("Nothing")
except:
    print("Nothing")

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + listOBJ[z].State + " | " + str(
        listOBJ[z].Latitude) + " | " + str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)
    sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].Type)

    wb_obj_w.save("Documents\OneTree_AUS_20052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\OneTree_AUS_20052021.xlsx")
from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json

my_path = "Documents\Australian_All_Suburb_13052021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row




my_path = "Documents\Liquorland_AUS_10062021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []

headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'cookie': '__uzma=3f11ce3c-4705-47a7-9d8f-6c13b96679e7; __uzmb=1623332420; AMCV_0B3D037254C7DE490A4C98A6%40AdobeOrg=1075005958%7CMCIDTS%7C18789%7CMCMID%7C58063441793552980883422418947050361035%7CMCAAMLH-1623937239%7C6%7CMCAAMB-1623937239%7CRKhpRz8krg2tLO6pguXWp5olkAcUniQYPHaMWWgdJ3xzPWQmdj0y%7CMCOPTOUT-1623339639s%7CNONE%7CvVersion%7C4.4.1; sqvisitor="id=febca045-edf0-4a5a-885a-e8cf1172459e"; _gcl_au=1.1.696874573.1623332445; _fbp=fb.2.1623332459653.902196686; _ga=GA1.3.1187729769.1623332465; rmStore=dmid:9145; stc115779=tsa:1623332463187.1687985600.070609.3778075385883113.1:20210610141117|env:1%7C20210711134103%7C20210610141117%7C2%7C1053756:20220610134117|uid:1623332463187.789531167.926867.115779.885942633.73:20220610134117|srchist:1053756%3A1%3A20210711134103:20220610134117; ins-storage-version=2; ORA_FPC=id=6e4ba75d-c3a8-4518-a780-49a3870985e5; WTPERSIST=; RT="r=https%3A%2F%2Fwww.liquorland.com.au%2Fstores&ul=1623341012943"; __uzmc=933565527481; __uzmd=1623430773; SSID=CABXqR1GAAAAAABVFsJgFwCAO0QWwmACAAAAAAAAAAAAdZbDYABuOssAAAMYHAAARBbCYAIAxgAAA6YbAABEFsJgAgDAAAADOSEAAEQWwmACAMoAAAPsGwAARBbCYAIA7AAAA8UiAABEFsJgAgA; SSSC=6.G6972159728449224727.2|192.8505:198.7078:202.7148:203.7192:236.8901; SSRT=dZbDYAADAA',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'
}

first = 2
last = 10

for i in range (first,last + 1):
    URL = 'https://www.liquorland.com.au/api/FindClosest/ll?lat='+str(postcode_sheet_obj.cell(row = i, column = 4).value)+'&lon='+str(postcode_sheet_obj.cell(row = i, column = 5).value)
    print(str(i) + " of " + str(last))
    print(URL)
    try:

        res = requests.request("GET", URL, headers=headers)

        if res.status_code == 200:
            output = json.loads(res.text)

            if len(output) > 0:
                for store in output:

                    try:

                        obj = OBJ()
                        obj.Title = str(store['storeName'])
                        obj.Address = str(store['address'])

                        obj.Suburb = str(store['suburb'])
                        obj.State = str(store['state'])
                        obj.Postcode = str(store['postcode'])
                        obj.Country = "Australia"
                        obj.Latitude = str(store['latitude'])
                        obj.Longitude = str(store['longitude'])

                        print(obj.Title + " | " + obj.Address + " | " + obj.Suburb + " | " + obj.State + " | " + str(
                            obj.Postcode) + " | " + obj.Country + " | " + str(obj.Latitude) + " | " + str(
                            obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                # print(str(obj.Title) + " " + str(listOBJ[z].Title))
                                if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                        listOBJ[i].Address) and str(obj.Suburb) == str(listOBJ[i].Suburb) and str(
                                    obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                                    listOBJ[i].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)
                    except:
                        continue
    except:
        # listError.append(URL)
        continue

    # break

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(str(j) + " of " +str(len(listOBJ)) +" | "+listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\Liquorland_AUS_10062021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\Liquorland_AUS_10062021.xlsx")
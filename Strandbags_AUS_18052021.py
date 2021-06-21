from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json
from types import SimpleNamespace

my_path = "Documents\Australian_All_Suburb_13052021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row
print("Total Row: " + str(postcode_max_row))




my_path = "Documents\Strandbags_AUS_18052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

# for i in range(2, 3170):
#     print(str(int(float(postcode_sheet_obj.cell(row = i, column = 2).value))))


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []

for i in range(1, 2650 + 1):
    URL = 'https://www.strandbags.com.au/store-locator/search?lat='+str(postcode_sheet_obj.cell(row = i, column = 4).value)+'&lng='+str(postcode_sheet_obj.cell(row = i, column = 5).value)
    print(URL)
    res = requests.request("GET", URL)

    if res.status_code == 200:
        # soup = BeautifulSoup(res.content, "html.parser")

        try:
            output = str(res.text).replace("\\n", "").replace("\\t", "").replace("  ", "")

            index1 = output.index('var stores =')
            index2 = output.index(';', index1)
            output = output[index1 + len('var stores ='):index2]
            zz = str(output).replace("\\n", "").replace("\\t", "").replace("  ", "").strip()

            locations = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

            try:

                for x in locations:
                    obj = OBJ()
                    obj.Title = str(x.attr_headline)
                    obj.Address = str(x.attr_address1)
                    obj.FullAddress = str(x.attr_address2) + ', ' + str(x.attr_address3)
                    obj.Suburb = str(x.attr_suburb)
                    obj.State = str(x.attr_state)
                    obj.Postcode = str(x.attr_postcode)
                    obj.Country = "Australia"
                    obj.Latitude = str(x.attr_lat)
                    obj.Longitude = str(x.attr_lng)

                    print(obj.Title + " | " + obj.Address + " | " + obj.Suburb + " | " + str(obj.Postcode)+ " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                    result = False

                    if len(listOBJ) > 0:
                        for z in range(len(listOBJ)):
                            if (str(obj.Title) == listOBJ[z].Title and str(obj.Address) == listOBJ[z].Address and str(
                                    obj.Postcode) == str(listOBJ[z].Postcode) and str(obj.Latitude) == str(listOBJ[z].Latitude) and str(obj.Longitude) == str(listOBJ[
                                z].Longitude)):
                                result = True
                                break

                    if result == False:
                        listOBJ.append(obj)

            except:
                continue

        except:
            listError.append(URL)
            continue

        break

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\Strandbags_AUS_18052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\Strandbags_AUS_18052021.xlsx")
from bs4 import BeautifulSoup
import requests
import openpyxl
import requests 

import json
from types import SimpleNamespace

my_path = "Documents/NZ_Cities_11022021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row

my_path = "Documents/KMART_NZ_12062021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
listError = []


URL = 'https://www.kmart.co.nz/webapp/wcs/stores/servlet/AjaxStoreLocatorMapResultsView'
headers = {
                # ':authority': 'www.kmart.co.nz',
                # ':method': 'POST',
                # ':path': '/webapp/wcs/stores/servlet/AjaxStoreLocatorMapResultsView',
                # ':scheme': 'https',
                'accept': '*/*',
                'accept-encoding': 'gzip, deflate, br',
                'accept-language': 'en-US,en;q=0.9',
                'content-length': '66',
                'content-type': 'application/x-www-form-urlencoded',
                'cookie': 'optimizelyEndUserId=JfBzWCxPCY6rS5FOJ_1623393679900; WC_PERSISTENT=JjBqeANR%2FJk25enK5Qtwu%2Bp%2FALI%3D%0A%3B2021-06-11+16%3A41%3A19.901_1623393679899-105959_20701_-1002%2C-1%2CNZD_20701; visid_incap_1039273=OhGwd1S1RtGtm2R0Yh7re48Fw2AAAAAAQUIPAAAAAACqyTzKfed+cmZIDzmYgFEz; optimizelyEndUserId=JfBzWCxPCY6rS5FOJ_1623393679900; cus_adl_state=NI; _fbp=fb.2.1623393743065.1366546158; _ga=GA1.3.794998268.1623393749; _gid=GA1.3.1907371005.1623393749; usrLS=Fri%20Jun%2011%202021%2012%3A13%3A18%20GMT%2B0530%20(India%20Standard%20Time); JSESSIONID=0000J-aImKjcbYWJE0Io_umcUXr:1bu08ukhe; KMUSR=G||false|-1002|; WC_SESSION_ESTABLISHED=true; WC_AUTHENTICATION_-1002=-1002%2CsXdG2OlftFaNqlDzZ1KxZB71p6I%3D; WC_ACTIVEPOINTER=-1%2C20701; WC_USERACTIVITY_-1002=-1002%2C20701%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2CS78z7uPbXiY2tQogBovgreT8o6V4LDAxN6t2G6kl1TkNWPEwOMJnzcfCfwIaZK9l0pPW7wp%2FwPUEgDMblCtNyvLHFTW2YNyyGEvYVF2h3WJAkmuoVI9KVEVGM0RGxidkUxLz2lQh2xRhUptQFKlI8udEHPB13iFXTjsBrM0kn34%2B2OWY97H69GoOpfqDU7RiobwN73qgdZFOSaUCamtYPA%3D%3D; WC_GENERIC_ACTIVITYDATA=[22729831535%3Atrue%3Afalse%3A0%3Ah3CjDMObfAGpZa4GTmryEaJfSso%3D][com.ibm.commerce.context.audit.AuditContext|1623393679899-105959][com.ibm.commerce.store.facade.server.context.StoreGeoCodeContext|null%26null%26null%26null%26null%26null][CTXSETNAME|Store][com.ibm.commerce.context.globalization.GlobalizationContext|-1%26NZD%26-1%26NZD][com.ibm.commerce.catalog.businesscontext.CatalogContext|20102%26null%26false%26false%26false][com.ibm.commerce.context.ExternalCartContext|null][com.ibm.commerce.context.base.BaseContext|20701%26-1002%26-1002%26-1][com.ibm.commerce.context.experiment.ExperimentContext|null][com.ibm.commerce.context.entitlement.EntitlementContext|4000000000000001507%264000000000000001507%26null%26-2000%26null%26null%26null][com.ibm.commerce.giftcenter.context.GiftCenterContext|null%26null%26null]; incap_ses_364_1039273=z0p2fQvAG2+CTDt0aDANBfMwxGAAAAAACL8LZexDr4Svm1ZwNd3FhQ==; _gat_UA-7745282-3=1; iter_id=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJhaWQiOiI2MGMzMDVkNWQyYmQ1NzAwMDE3MzFiZWUiLCJjb21wYW55X2lkIjoiNWVlMjUyNGQ4ZGEyMTYwMDAxMWU2NTRmIiwiaWF0IjoxNjIzNDcwMzM1fQ.Ku80jlR_QG2LtsyL2-cWIIk2RAjnALUD52hshBLX3EI; WC_stZip=2000; WC_stFind=2',
                'origin': 'https://www.kmart.co.nz',
                'referer': 'https://www.kmart.co.nz/store-locator',
                'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="91", "Chromium";v="91"',
                'sec-ch-ua-mobile': '?0',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-origin',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36',
                'x-requested-with': 'XMLHttpRequest'

            }

try:
        for i in range(3, postcode_max_row+1):
            Data = {
                'geoCodeLatitude': str(postcode_sheet_obj.cell(row = i, column = 2).value),
                'geoCodeLongitude': str(postcode_sheet_obj.cell(row = i, column = 1).value),
                'cityId': '-888'
            }
            print(Data)
            print(i)

                

            response = requests.post(URL , headers=headers , data=Data)

            if response.status_code == 200:
                zz = str(response.text).replace("\\n", "").replace("\n", "").replace("\\t", "").replace("\t", "").replace("  ", "").replace("\'", "\"").strip()

                output = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

                for y in range(len(output['physicalStore'])):
                    try:
                        x = output['physicalStore'][y]

                        obj = OBJ()
                        obj.Title = str(x['htmlIdentifier'])
                        obj.Address = str(x['htmlAddress1'])

                        

                    
                        obj.Postcode = str(x['htmlAddress2'])
                        
                        obj.Latitude = str(x['latitude'])
                        obj.Longitude = str(x['longitude'])

                        print(obj.Title + " | " + obj.Address + " | "  + str(obj.Postcode) + " | "  + str(obj.Latitude) + " | " + str(obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for k in range(len(listOBJ)):
                                if (str(obj.Title) == str(listOBJ[k].Title) and str(obj.Address) == str(listOBJ[k].Address) and str(obj.Latitude) == str(listOBJ[k].Latitude) and str(obj.Longitude) == str(listOBJ[k].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)


                    except:
                        continue

except : 
    print("Exception: " + URL)
    listError.append(URL)    


j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" | "+listOBJ[z].Address +" | "+ str(listOBJ[z].Latitude) +" | "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    #sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].City)
    #sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].State)
    #sheet_obj_w.cell(row = j, column = 5).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row = j, column = 6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row = j, column = 7).value = str(str(listOBJ[z].Latitude))
    sheet_obj_w.cell(row = j, column = 8).value = str(str(listOBJ[z].Longitude))
    wb_obj_w.save("Documents/KMART_NZ_12062021.xlsx")

j = j + 10

if(len(listError) > 0):
    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/KMART_NZ_12062021.xlsx")
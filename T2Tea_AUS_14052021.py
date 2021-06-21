from bs4 import BeautifulSoup
import requests
import openpyxl
import requests 

import json
from types import SimpleNamespace


my_path = "Documents/T2Tea_AUS_14052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
listError = []

for z in range(0,10):

    url = "https://www.t2tea.com/on/demandware.store/Sites-UNI-T2-APAC-Site/en_AU/Stores-FindStores?showMap=false&postalCode="+str(z)

    try:

        print(url)

        response = requests.get(url).json()

        listLocation = response["stores"]

        for x in listLocation:

            obj = OBJ()
            obj.Title = str(x["name"])
            obj.Address = str(x["address1"])
            obj.FullAddress = str(x["address2"])
            obj.City = str(x["city"])
            obj.State = str(x["stateCode"])
            obj.Postcode = str(x["postalCode"])
            obj.Latitude = str(x["latitude"])
            obj.Longitude = str(x["longitude"])

            obj.Country = "Australia"


            print(obj.Title + " | " + obj.Address + " | " + obj.City + " | " + obj.State + " | " + str(obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

            result = False

            if len(listOBJ) > 0:
                for i in range(len(listOBJ)):
                    if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Postcode) == str(listOBJ[i].Postcode) and str(obj.Address) == str(listOBJ[i].Address) and str(obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(listOBJ[i].Longitude)):
                        result = True
                        break

            if result == False:
                listOBJ.append(obj)

    except : 
        print("Exception: " + url)
        listError.append(url)
        continue

    # break


j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" "+listOBJ[z].Address +" "+ str(listOBJ[z].Latitude) +" "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row = j, column = 5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row = j, column = 6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row = j, column = 7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row = j, column = 8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row = j, column = 9).value = str(listOBJ[z].Longitude)
    wb_obj_w.save("Documents/T2Tea_AUS_14052021.xlsx")

j = j + 10

if (len(listError) > 0):
    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/T2Tea_AUS_14052021.xlsx")
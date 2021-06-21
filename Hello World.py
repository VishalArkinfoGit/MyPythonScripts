
from bs4 import BeautifulSoup
import requests, json
import openpyxl

# my_path = "Documents\Australian_ByName_Postcodes_12022021.xlsx"
# postcode_wb_obj = openpyxl.load_workbook(my_path)
# postcode_sheet_obj = postcode_wb_obj.active
# postcode_max_col = postcode_sheet_obj.max_column
# postcode_max_row = postcode_sheet_obj.max_row




class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
obj = OBJ()
listOBJ.append(obj)

for i in range(2, 898+1):
    print(i)
    # URL = 'https://agents.helloworld.com.au/search-location/' + str(postcode_sheet_obj.cell(row=i, column=1).value)
    URL = 'https://agents.helloworld.com.au/searchLocation/searchResultsListByLocation'
    headers = {
        'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'cookie': 'timezoneoffset=-330; ARRAffinity=a1fafedc8aa87fe4f9a92b5f7e3391af401b5801e8e94f5c9bf0be460ef37d3c; ARRAffinitySameSite=a1fafedc8aa87fe4f9a92b5f7e3391af401b5801e8e94f5c9bf0be460ef37d3c; visid_incap_1691371=EtvE9wOSTTu1nE91Yyk1+RpxsmAAAAAAQUIPAAAAAACs73FmbJX/vMO0V6B82w/C; nlbi_1691371=o7rSTKP2fhpGmLLNMeQQdAAAAABjlhMWJx3fgIINjczVrFWQ; incap_ses_477_1691371=SxrkQZjQAGJEroaEWKWeBhtxsmAAAAAA64lE0d5tWHtxMQRGfk6Qsg==; optimizelyEndUserId=oeu1622307095156r0.8013922480882691; _gcl_au=1.1.21405913.1622307133; _ga=GA1.3.1694202104.1622307144; _gid=GA1.3.1827773322.1622307144; _dc_gtm_UA-42793227-3=1; _gat=1; __utma=72248323.1694202104.1622307144.1622307145.1622307145.1; __utmc=72248323; __utmz=72248323.1622307145.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); __utmt_UA-42793227-2=1; _fbp=fb.2.1622307150407.484872668; _ga=GA1.4.1694202104.1622307144; _gid=GA1.4.1827773322.1622307144; _gat_UA-42793227-2=1; visid_incap_1602118=U3h93HLpSsGuyIoDx5JYw1ZxsmAAAAAAQUIPAAAAAABkSC9TLTANzbHEVIViSscp; nlbi_1602118=kvq5DyWTSUFWtqLA6JfP+QAAAADGa/Y9hMaLqEvJKVc6hE88; incap_ses_477_1602118=wJXLBhSjHFHx1IaEWKWeBldxsmAAAAAAtGHrFq5nqxBejD9DVtznmg==; timezoneoffset=-330; __utmb=72248323.3.10.1622307145',
        'dnt': '1',
        'origin': 'https://agents.helloworld.com.au',
        'referer': 'https://agents.helloworld.com.au/',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
        'sec-ch-ua-mobile': '?0',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36',
        'x-requested-with': 'XMLHttpRequest'
    }

    data ={
        "searchLocation": '2000'
    }
    res = requests.post(URL, headers=headers, data=data)
    print(res)
    if res.status_code == 200:
            output = json.loads(res.text)

            if len(output['Agents']) > 0:
                for store in output['Agents']:

                    try:

                        obj = OBJ()
                        obj.Title = str(store['Name'])
                        obj.Address = str(store['ContactInfo']['FullAddress'])
                        obj.FullAddress = str(store['address2'])
                        obj.Latitude = str(store['Geolocation']['Latitude'])
                        obj.Longitude = str(store['Geolocation']['Longitude'])

                        print(obj.Title + " | " + obj.City + " | " + obj.State + " | " + str(
                            obj.Postcode) + " | " + obj.Country + " | " + str(obj.Latitude) + " | " + str(
                            obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                # print(str(obj.Title) + " " + str(listOBJ[z].Title))
                                if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                        listOBJ[i].Address) and str(obj.City) == str(listOBJ[i].City) and str(
                                    obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                                    listOBJ[i].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)
                    except:
                        continue

my_path = "Documents\HelloWorld_June.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

j=0
print(len(listOBJ))
for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" "+listOBJ[z].Postcode +" "+listOBJ[z].Suburb +" "+listOBJ[z].State+" "+listOBJ[z].Country +" "+ str(listOBJ[z].Latitude) +" "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row = j, column = 5).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row = j, column = 6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row = j, column = 7).value = str(str(listOBJ[z].Latitude))
    sheet_obj_w.cell(row = j, column = 8).value = str(str(listOBJ[z].Longitude))
    wb_obj_w.save("C:/Python/Documents/HelloWorld_June.xlsx")

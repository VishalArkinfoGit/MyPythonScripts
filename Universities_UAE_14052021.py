from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime
import string

now = datetime.now()

my_path = "Documents/Universities_UAE_14052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""

listOBJ = []
listError = []

dt_string = now.strftime("%d%m%Y")

listLink = []



for z in range(1, 7):
    url = "https://www.edarabia.com/universities/uae/?pg=" + str(z)
    print(url)
    listError.append(url)
    try:

        res = requests.request("GET", url)

        if res.status_code == 200:
            soup = BeautifulSoup(res.content, "html.parser")

            try:
                list = soup.find('div', class_="adv-filter-content").find("div", class_="content-box-0")\
                    .find_all('div', class_='list-items')

                for li in list:
                    try:
                        link = li.find("a", class_="red-button")['href']

                        print(str(link))

                        result = False

                        if len(listLink) > 0:
                            for i in range(len(listOBJ)):
                                if (str(link) == str(listLink[i])):
                                    result = True
                                    break

                        if result == False:
                            listLink.append(link)

                    except:
                        continue

            except:
                listError.append(url)
                continue

        # break

    except:
        print("*******************************************")
        print("Location Not Found: " + url)
        listError.append(url)
        print("*******************************************")
        print("\n")
        continue


for url in listLink:

    print(url)

    try:

        res = requests.request("GET", url)

        if res.status_code == 200:
            soup = BeautifulSoup(res.content, "html.parser")

            try:
                obj = OBJ()

                obj.Title = soup.find('div', class_="single-top").find("h1").text

                try:
                    ul = soup.find("ul", class_="list-items-top").find_all('li')

                    for li in ul:
                        try:
                            if ("Address" in str(li.find("span").text)):

                                try:

                                    output = str(li.prettify()).replace('\n','').replace('  ','').strip()

                                    try:

                                        index1 = output.index('</span>')
                                        index2 = output.index('<a', index1)
                                        obj.Address = output[index1 + len('</span>'):index2]
                                    except:
                                        obj.Address = ''


                                    try:
                                        index3 = output.index('<a', index2)
                                        index3 = output.index('>', index3)
                                        index4 = output.index('<', index3)
                                        obj.State = output[index3 + len('>'):index4]
                                    except:
                                        obj.State = ''


                                    try:
                                        index3 = output.index('<a', index4)
                                        index3 = output.index('>', index3)
                                        index4 = output.index('<', index3)
                                        obj.Country = output[index3 + len('>'):index4]
                                    except:
                                        obj.Country = ''



                                    try:
                                        index3 = output.index('location_lat', index4)
                                        index3 = output.index('=', index3)
                                        index4 = output.index(';', index3)
                                        obj.Latitude = str(output[index3 + len('>'):index4]).replace('\'','')
                                    except:
                                        obj.Latitude = ''



                                    try:
                                        index3 = output.index('location_lng', index4)
                                        index3 = output.index('=', index3)
                                        index4 = output.index(';', index3)
                                        obj.Longitude = str(output[index3 + len('>'):index4]).replace('\'','')
                                    except:
                                        obj.Longitude = ''


                                except:
                                    continue

                            else:
                                continue

                        except:
                            continue

                except:
                    obj.Address = ''
                    obj.State = ''
                    obj.Latitude = ''
                    obj.Longitude = ''

                print(obj.Title + " | " + obj.Address + " | " + str(obj.State) + " | " + str(
                    obj.Latitude) + " | " + str(obj.Longitude))

                result = False

                if len(listOBJ) > 0:
                    for i in range(len(listOBJ)):
                        if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                listOBJ[i].Address) and str(obj.State) == str(listOBJ[i].State) and str(
                                obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                                listOBJ[i].Longitude)):
                            result = True
                            break

                if result == False:
                    listOBJ.append(obj)

            except:
                listError.append(url)
                continue

        # break

    except:
        print("*******************************************")
        print("Location Not Found: " + url)
        listError.append(url)
        print("*******************************************")
        print("\n")
        continue

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + listOBJ[z].State + " | " + str(
        listOBJ[z].Latitude) + " | " + str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)
    sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].Type)

    wb_obj_w.save("Documents/Universities_UAE_14052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/Universities_UAE_14052021.xlsx")
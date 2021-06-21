from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime

my_path = "Documents\Amcal_AUS_14052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active



class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []



def recursive(list, html, startIndex):

    if(html.find('var test1 =', startIndex) > -1):

        obj = OBJ()

        index1 = html.find('var test1 =', startIndex)
        index1 = html.find('=', index1)
        index2 = html.find(';', index1)
        strJson = html[index1 + 1: index2]

        try:
            data = json.loads(strJson)

            try: obj.Title = data['Description'][0]['displayStoreName']
            except: obj.Title = ""

            try: obj.City = data['city']
            except: obj.City = ""

            try:
                for address in data['addressLine']:
                    obj.Address = obj.Address + address + ","

            except: obj.Address = ""

            try: obj.State = data['stateOrProvinceName']
            except: obj.State = ""

            try: obj.Postcode = data['postalCode']
            except: obj.Postcode = ""

            try: obj.Latitude = data['latitude']
            except: obj.Latitude = ""

            try: obj.Longitude = data['longitude']
            except: obj.Longitude = ""
        except:
            obj = None

        if obj != None:
            print(obj.Title + " | " + obj.Address + " | " + obj.City + " | " + obj.State + " | " + str(obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(
                obj.Longitude))

            result = False

            if len(list) > 0:
                for i in range(len(list)):
                    if (str(obj.Title) == str(list[i].Title) and str(obj.Address) == str(list[i].Address) and str(
                            obj.Postcode) == str(list[i].Postcode) and str(obj.Latitude) == str(
                            list[i].Latitude) and str(obj.Longitude) == str(list[i].Longitude)):
                        result = True
                        break

            if result == False:
                list.append(obj)

        if (html.find('var test1 =', index2) > -1):
            return recursive(list,html, index2)

        return list

    else: return list


listRegien = ['vic', 'nsw', 'qld', 'tas', 'act', 'wa', 'sa', 'nt']

for reg in listRegien:
    url = "https://www.amcal.com.au/store-locator/" + reg
    print(url)
    res = requests.request("GET", url)

    if res.status_code == 200:
        soup = BeautifulSoup(res.content, "html.parser")

        try:
            output = str(soup.prettify()).replace('\n','').replace('\r','').replace('\t','').replace('  ','')

            recursive(listOBJ, output,0)

        except:
            listError.append(url)
            continue

    # break



j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + listOBJ[z].City + " | " + listOBJ[z].State + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents/Amcal_AUS_14052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/Amcal_AUS_14052021.xlsx")
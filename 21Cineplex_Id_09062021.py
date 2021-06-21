from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime

my_path = "Documents/21Cineplex_Id_09062021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active



class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""


listOBJ = []
listCities = []
listError = []

url = 'https://21cineplex.com/theaters'

headers = {'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9','cookie': 'PHPSESSID=80shi7v6gc4kifh0q1k4uhrc6g; __utma=117930442.1829253272.1623253912.1623253912.1623253912.1; __utmc=117930442; __utmz=117930442.1623253912.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); kota=77; __utmb=117930442.10.10.1623253912'}

res = requests.get(url, headers=headers)

if res.status_code == 200:
    soup = BeautifulSoup(res.content, "html.parser")
    output = soup.find('select', id="cityChanged").find_all('option')
    for div in output:
        print(div['value'])

# listRegien = ['ACT', 'NSW', 'NT', 'QLD', 'SA', 'TAS', 'VIC', 'WA']
#
# url = "https://6fb4630d53ba4184a1e9bf9f861dc00f.ap-southeast-2.aws.found.io/fcl_store_prod/_search"
# headers = {
#   'accept': 'application/json',
#   'authorization': 'Basic ZmNhdV93ZWJfY3VzdG9tZXI6WmpXYkN1Q19Vck9MVFJqSg==',
#   'Content-Type': 'application/json'
# }
#
# for reg in listRegien:
#     print(url +"/" + reg)
#     payload = "{\"query\":{\"bool\":{\"must\":{\"match\":{\"state\":\""+reg+"\"}}}},\"size\":9999}"
#     res = requests.request("POST", url, headers=headers, data=payload)
#
#     if res.status_code == 200:
#
#         try:
#             output = json.loads(res.text)
#
#             locations = output['hits']['hits']
#
#             for x in locations:
#                 try:
#                     obj = OBJ()
#                     obj.Title = str(x['_source']['name'])
#                     obj.Address = str(x['_source']['address1'])
#                     try: obj.FullAddress = str(x['_source']['address2'])
#                     except: obj.FullAddress = ""
#                     obj.City = str(x['_source']['locality'])
#                     obj.State = str(x['_source']['state'])
#                     obj.Postcode = str(x['_source']['postcode'])
#                     obj.Type = str(x['_source']['type'])
#                     obj.Country = str(x['_source']['country_name'])
#                     obj.Latitude = str(x['_source']['geo_location']['lat'])
#                     obj.Longitude = str(x['_source']['geo_location']['lon'])
#
#                     print(obj.Title + " | " + obj.Address + " | " + obj.City + " | " + str(obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))
#
#                     result = False
#
#                     if len(listOBJ) > 0:
#                         for z in range(len(listOBJ)):
#                             if (str(obj.Title) == listOBJ[z].Title and str(obj.Postcode) == str(listOBJ[z].Postcode) and str(obj.Latitude) == str(listOBJ[z].Latitude) and str(obj.Longitude) == str(listOBJ[z].Longitude)):
#                                 result = True
#                                 break
#
#                     if result == False:
#                         listOBJ.append(obj)
#                 except:
#                     continue
#         except:
#             listError.append(url)
#             continue
#
#     # break
#
#
#
# j = 0
#
# for z in range(len(listOBJ)):
#     j = j + 1
#     print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(
#         listOBJ[z].Latitude) + " | " + str(
#         listOBJ[z].Longitude))
#     sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
#     sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
#     sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
#     sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
#     sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
#     sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
#     sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
#     sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
#     sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)
#     sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].Type)
#
#     wb_obj_w.save("Documents/21Cineplex_Id_09062021.xlsx")
#
# j = j + 10
#
# if (len(listError) > 0):
#
#     sheet_obj_w.cell(row=j, column=1).value = 'ERROR'
#
#     for z in listError:
#         j = j + 1
#         sheet_obj_w.cell(row=j, column=1).value = str(z)
#         wb_obj_w.save("Documents/21Cineplex_Id_09062021.xlsx")
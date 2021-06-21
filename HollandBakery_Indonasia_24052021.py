from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json


my_path = "Documents\HollandBakery_Indonasia_24052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []

headers = {
  'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Cookie': 'www_pyrocms=a%3A4%3A%7Bs%3A10%3A%22session_id%22%3Bs%3A32%3A%226dce9128f35ce2b62ce208e983095ecc%22%3Bs%3A10%3A%22ip_address%22%3Bs%3A13%3A%22103.77.235.42%22%3Bs%3A10%3A%22user_agent%22%3Bs%3A114%3A%22Mozilla%2F5.0+%28Windows+NT+6.3%3B+Win64%3B+x64%29+AppleWebKit%2F537.36+%28KHTML%2C+like+Gecko%29+Chrome%2F90.0.4430.212+Safari%2F537.36%22%3Bs%3A13%3A%22last_activity%22%3Bi%3A1621854472%3B%7Ddde5b3e94d59eaeb56fdaeb446ccc4f8; PHPSESSID=1b2uss6ns5hv76gdsdv85169j0; _ga=GA1.3.1795518019.1621879681; _gid=GA1.3.1500949078.1621879681; __tawkuuid=e::hollandbakery.co.id::MdaveKRfMlot0pfhkPpseDobGHK1c6UvY2bVHuk0wu0zxiOEgREwbnBb7lBxtG3T::2; TawkConnectionTime=0',
    'DNT': '1',
    'Host': 'www.hollandbakery.co.id',
    'Origin': 'https://www.hollandbakery.co.id',
    'Referer': 'https://www.hollandbakery.co.id/grid-view/5?',
    'X-Requested-With': 'XMLHttpRequest'
}

for i in range(1, 2):
    URL = 'https://www.hollandbakery.co.id/get-group'
    print(URL)
    try:

        res = requests.request("POST", URL, headers=headers)

        if res.status_code == 200:
            z = json.loads(res.text)

            if len(z) > 0:
                for j in z:
                    for store in j['detail_toko']:

                        try:

                            obj = OBJ()
                            obj.Title = str(store['name'])
                            obj.Address = str(store['alamat'])

                            obj.City = str(store['provinsi'])
                            obj.State = str(store['kota'])
                            obj.Country = "Australia"
                            obj.Latitude = str(store['latitude'])
                            obj.Longitude = str(store['longitude'])

                            print(obj.Title + " | " + obj.Address + " | " + obj.City + " | " + obj.State + " | " + obj.Country + " | " + str(obj.Latitude) + " | " + str(
                                obj.Longitude))

                            result = False

                            if len(listOBJ) > 0:
                                for i in range(len(listOBJ)):
                                    # print(str(obj.Title) + " " + str(listOBJ[z].Title))
                                    if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                            listOBJ[i].Address) and str(obj.City) == str(listOBJ[i].City) and str(
                                        obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                                        listOBJ[i].Longitude)):
                                        result = True
                                        break

                            if result == False:
                                listOBJ.append(obj)
                        except:
                            continue
    except:
        # listError.append(URL)
        continue

    break

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\HollandBakery_Indonasia_24052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\HollandBakery_Indonasia_24052021.xlsx")
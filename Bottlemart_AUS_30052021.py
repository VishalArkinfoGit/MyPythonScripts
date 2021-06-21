from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json
from types import SimpleNamespace
import math
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")

my_path = "Documents\Australia_Cities_30052021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row

# for i in range(170, postcode_max_row+1):
#     # R = 6371;
#     # lat1 = float(str(postcode_sheet_obj.cell(row = i, column = 2).value))
#     # lat2 = float(str(postcode_sheet_obj.cell(row = i+1, column = 2).value))
#     # lon1 = float(str(postcode_sheet_obj.cell(row = i, column = 3).value))
#     # lon2 = float(str(postcode_sheet_obj.cell(row = i+1, column = 3).value))
#     # dLat = deg2rad(lat2 - lat1)
#     # dLon = deg2rad(lon2 - lon1)
#     # a = math.sin(dLat / 2) * math.sin(dLat / 2) + math.cos(deg2rad(lat1)) * math.cos(deg2rad(lat2)) * math.sin(dLon / 2) * math.sin(dLon / 2)
#     # c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
#     # d = R * c
#     #
#     # print(str(lat1) +' - '+str(lon1) +' | '+str(lat2) +' - '+str(lon2) +' | '+str(d))
#
#     try:
#         g = geolocator.geocode(str(postcode_sheet_obj.cell(row = i, column = 1).value) + ' ' + str(postcode_sheet_obj.cell(row = i, column = 2).value) + ' Australia')
#         postcode_sheet_obj.cell(row=i, column=3).value = str(g.latitude)
#         postcode_sheet_obj.cell(row=i, column=4).value = str(g.longitude)
#
#         lst = str(g.address).split(',')
#         postcode = lst[len(lst)-2]
#         postcode_sheet_obj.cell(row=i, column=5).value = str(postcode)
#
#         postcode_wb_obj.save("Documents\Australia_Cities_30052021.xlsx")
#     except:
#         continue





my_path = "Documents\Bottlemart_AUS_30052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []

for i in range(1, 4 + 1):
    URL = 'https://app.ehoundplatform.com/api/1.3/proximity_search?output=json&api_key=1qt7ts2bsw8a73v&lat='+str(postcode_sheet_obj.cell(row = i, column = 4).value)+'&lon='+str(postcode_sheet_obj.cell(row = i, column = 5).value)+'&count=99&max_distance=999'
    print(URL)
    res = requests.request("GET", URL)

    if res.status_code == 200:

        try:
            output = res.json()

            try:

                locations = output['record_set']

                for x in locations:
                    obj = OBJ()
                    obj.Title = str(x['account_name'])
                    try:
                        if(str(x['address']['shop_unit']) != ''):
                            obj.Address = str(x['address']['shop_unit']) + ',' + str(x['address']['street_address'])
                        else:
                            obj.Address = str(x['address']['street_address'])
                    except:
                        obj.Address = str(x['address']['street_address'])
                    obj.City = str(x['address']['town'])
                    obj.State = str(x['address']['state'])
                    obj.Postcode = str(x['address']['postal_code'])
                    obj.Country = "Australia"
                    obj.Latitude = str(x['latitude'])
                    obj.Longitude = str(x['longitude'])

                    print(obj.Title + " | " + obj.Address + " | " + obj.City + " | " + str(obj.Postcode)+ " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                    result = False

                    if len(listOBJ) > 0:
                        for z in range(len(listOBJ)):
                            if (str(obj.Title) == listOBJ[z].Title and str(obj.Address) == listOBJ[z].Address and str(
                                    obj.Postcode) == str(listOBJ[z].Postcode) and str(obj.Latitude) == str(listOBJ[z].Latitude) and str(obj.Longitude) == str(listOBJ[z].Longitude)):
                                result = True
                                break

                    if result == False:
                        listOBJ.append(obj)

            except:
                continue

        except:
            listError.append(URL)
            continue

        break

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\Bottlemart_AUS_30052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\Bottlemart_AUS_30052021.xlsx")
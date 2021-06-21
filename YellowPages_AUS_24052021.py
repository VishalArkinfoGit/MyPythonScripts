from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json
from types import SimpleNamespace

my_path = "Documents\YellowPages_AUS_24052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

# for i in range(2, 3170):
#     print(str(int(float(postcode_sheet_obj.cell(row = i, column = 2).value))))


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""


listOBJ = []
listError = []

headers = {
  'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
  'accept-encoding': 'gzip, deflate, br',
  'accept-language': 'en-US,en;q=0.9',
  'cache-control': 'max-age=0',
  'cookie': '_vwo_uuid_v2=D308924CE0C2A9CC17A9456C4557D61D5|8dbaa0ff82209af2ef672e4a6ba8326a; _vwo_uuid=D308924CE0C2A9CC17A9456C4557D61D5; s_ecid=MCMID|64569858247329101162770077466101935047; _wingify_pc_uuid=c057eecea3a1460da071b3266c2a3885; wingify_donot_track_actions=0; BVBRANDID=e6daa803-07c9-4a8f-8f76-55b101625a2d; _vis_opt_exp_217_combi=1; locationClue=Victoria; _vwo_ds=3:a_0,t_0:-1$1621612125:94.4529111:::3_0,2_0:0; RT="z=1&dm=www.yellowpages.com.au&si=14f886ca-6ab9-4832-8463-a444bbd3ba43&ss=kp4bhg0m&sl=0&tt=0"; yellow-guid=3f02bd27-a6be-4c13-b9ea-a47305e71f97; _vis_opt_s=5|; _vis_opt_test_cookie=1; _vwo_sn=438770:1; AMCVS_8412403D53AC3D7E0A490D4C@AdobeOrg=1; AMCV_8412403D53AC3D7E0A490D4C@AdobeOrg=-1124106680|MCIDTS|18774|MCMID|64569858247329101162770077466101935047|MCAAMLH-1622655692|6|MCAAMB-1622655692|RKhpRz8krg2tLO6pguXWp5olkAcUniQYPHaMWWgdJ3xzPWQmdj0y|MCOPTOUT-1622058092s|NONE|MCAID|NONE|vVersion|5.2.0; s_cc=true; bm_sv=3CCDB2109002F79CF3097664756A1B69~XgyAsX/SI/IxnSfKuc9jGe3KS3X7NSzrmxi+N9G7FMagb5T1RAzd/toUYoWSQW5DZubXQut8lqMWgFxuWseMJSC/xERLmctfa42w00KJluXtHYS1O2N6IHPjXAs/xp+wQazwulW1gXuA+lVG8NAPH8rQBRWLwhLINl1o5JeMnjg=; ak_bmsc=11532218C0F889376CD7DE9131A70F7858DD7004C73400004D88AE600F94C84C~plqc8OnD9LR35fgAmbRcQ+qZg1Jvjvsf0eGorX6jf/1XoijKhF/LFBKfH/ZrpAPj5U2y+uJCKqud2dWeAlGcRh6St+VePr5/F7o1gfEmYDE7yyXS+G4N+xZTByE293rKkDBGsi9WFgMH8CRsLHrnYt6B/d0e9HKN+TvxYC2r2MM0dTUnTpkDzTPfbLHGHuAtsaOOFGFp0U2xaeF4V6sPlSiFqjdEc9I9HeTyX4QFXit9aO5cn2qKU2P+c7RUBgrdaF',
  'dnt': '1',
  'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
  'sec-ch-ua-mobile': '?0',
  'sec-fetch-dest': 'document',
  'sec-fetch-mode': 'navigate',
  'sec-fetch-site': 'none',
  'sec-fetch-user': '?1',
  'upgrade-insecure-requests': '1',
  'user-agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'
}



def recursive(list, url):
    print(url)
    res = requests.request("GET", url, headers=headers)

    if res.status_code == 200:
        # soup = BeautifulSoup(res.content, "html.parser")

        try:
            output = str(res.text).replace("\\n", "").replace("\n", "").replace("\\t", "").replace("\t", "").replace(
                "  ", "")

            index1 = output.index('__INITIAL_STATE__')
            index1 = output.index('{', index1)
            index2 = output.index('</script>', index1)
            output = output[index1:index2 - 1]
            zz = str(output).replace("\\n", "").replace("\\t", "").replace("  ", "").strip()

            locations = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

            try:
                locations = locations.model.inAreaResultViews

                for x in locations:
                    try:
                        obj = OBJ()
                        obj.Title = str(x.name)
                        obj.Address = str(x.addressView.addressLine)
                        obj.FullAddress = str(x.addressView.asContactCardFormat)
                        obj.Suburb = str(x.addressView.suburb)
                        obj.State = str(x.addressView.state)
                        obj.Postcode = str(x.addressView.postCode)
                        obj.Country = "Australia"
                        obj.Latitude = str(x.addressView.latitude)
                        obj.Longitude = str(x.addressView.longitude)
                        obj.Type = str(x.category.name)

                        print(obj.Title + " | " + obj.Address + " | " + obj.Suburb + " | " + str(
                            obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                        result = False

                        if len(list) > 0:
                            for z in range(len(list)):
                                if (str(obj.Title) == list[z].Title and str(obj.Address) == list[
                                    z].Address and str(
                                    obj.Postcode) == str(list[z].Postcode) and str(obj.Latitude) == str(
                                    list[z].Latitude) and str(obj.Longitude) == str(list[
                                                                                           z].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            list.append(obj)
                    except:
                        continue
            except:
                print('Error')

        except:
            print(url)

    return list

listState = [
    # 'Australian Capital Territory'
 # 'New South Wales'
 'Victoria'
# 'Queensland'
# 'South Australia',
# 'Western Australia',
# 'Tasmania',
# 'Northern Territory'
]


for state in listState:
    URL = 'https://www.yellowpages.com.au/search/listings?clue=Vets+%26+Veterinary+Surgeons&locationClue='+state+'&selectedViewMode=list'
    print(URL)
    res = requests.request("GET", URL, headers=headers)

    if res.status_code == 200:
        # soup = BeautifulSoup(res.content, "html.parser")

        try:
            output = str(res.text).replace("\\n", "").replace("\n", "").replace("\\t", "").replace("\t", "").replace("  ", "")

            index1 = output.index('__INITIAL_STATE__')
            index1 = output.index('{', index1)
            index2 = output.index('</script>', index1)
            output = output[index1:index2-1]

            zz = str(output).replace("\\n", "").replace("\\t", "").replace("  ", "").strip()

            locations = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

            try:

                totalResults = locations.model.pagination.totalResults

                searchResultsPerPage = locations.model.pagination.searchResultsPerPage

                totalPages = int(totalResults / searchResultsPerPage) + 1

                for no in range(1,totalPages+1):
                    listOBJ.extend(recursive(listOBJ, URL + '&pageNumber='+str(no)))
                    # if no == 28:
                    #     break

            except:
                continue

        except:
            listError.append(URL)
            continue

        # break

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)
    sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].Type)

    wb_obj_w.save("Documents\YellowPages_AUS_24052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\YellowPages_AUS_24052021.xlsx")
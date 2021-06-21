import urllib.parse
import openpyxl
import requests
import json
from bs4 import BeautifulSoup

my_path = "Documents\AUS_Location_LatLong_14062021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row




my_path = "Documents\OfficeHub_AUS_10062021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Url = ""
    PropertyId = ""


listOBJ = []
listError = []
listLink = []

last = 5

for i in range(2, last-1):
    URL = 'https://d2a2bnk14z5yo9.cloudfront.net/api/search'

    try:
        params = {'location': str(postcode_sheet_obj.cell(row = i, column = 1).value)+' '+str(postcode_sheet_obj.cell(row = i, column = 2).value)+' Australia',
                  'latitude': str(postcode_sheet_obj.cell(row = i, column = 5).value),
                  'longitude': str(postcode_sheet_obj.cell(row = i, column = 6).value),
                  'radius': '999999',
                  'perPage': '48'}

        # params = {'latitude': '-33.8688197',
        #           'longitude': '151.2092955',
        #           'radius': '999999',
        #           'perPage': '48'}

        URL = 'https://d2a2bnk14z5yo9.cloudfront.net/api/search?'+urllib.parse.urlencode(params)
        print(str(i) + " of " +str(last) +" | "+ URL)
        res = requests.request("GET", URL)

        if res.status_code == 200:
            output = json.loads(res.text)

            totalPages = output['totalPages']

            if totalPages > 1:

                output = output['data']

                if len(output) > 0:
                    a = 0
                    for store in output:
                        a = a + 1
                        try:

                            obj = OBJ()
                            obj.Title = str(store['name'])
                            obj.Address = str(store['address'])

                            obj.City = str(store['city'])
                            obj.State = str(store['state'])
                            obj.Postcode = str(store['zipcode'])
                            obj.Country = "Australia"
                            obj.Latitude = str(store['geolocationLat'])
                            obj.Longitude = str(store['geolocationLong'])
                            obj.Url = str(store['urlPrefix'])
                            obj.PropertyId = str(store['propertyId'])

                            print(obj.Title + " | " + obj.City + " | " + obj.State + " | " + str(
                                obj.Postcode) + " | " + obj.Country + " | " + str(
                                obj.Latitude) + " | " + str(
                                obj.Longitude))

                            result = False

                            if len(listLink) > 0:
                                for i in range(len(listLink)):
                                    # print(str(obj.Title) + " " + str(listLink[z].Title))
                                    if (str(obj.Url) == str(listLink[i].Url) and str(obj.Latitude) == str(listLink[i].Latitude) and str(obj.Longitude) == str(listLink[i].Longitude)):
                                        result = True
                                        break

                            if result == False:
                                listLink.append(obj)
                        except:
                            continue

                        break

                for no in range(2, totalPages+1):

                    print(str(i) + " of " +str(last) +" | "+ URL + '&page='+str(no))
                    try:

                        res = requests.request("GET", URL + '&page='+str(no))

                        if res.status_code == 200:
                            output = json.loads(res.text)

                            output = output['data']

                            if len(output) > 0:
                                a = 0
                                for store in output:
                                    a = a + 1
                                    try:

                                        obj = OBJ()
                                        obj.Title = str(store['name'])
                                        obj.Address = str(store['address'])

                                        obj.City = str(store['city'])
                                        obj.State = str(store['state'])
                                        obj.Postcode = str(store['zipcode'])
                                        obj.Country = "Australia"
                                        obj.Latitude = str(store['geolocationLat'])
                                        obj.Longitude = str(store['geolocationLong'])
                                        obj.Url = str(store['urlPrefix'])
                                        obj.PropertyId = str(store['propertyId'])

                                        print(obj.Title + " | " + obj.City + " | " + obj.State + " | " + str(
                                            obj.Postcode) + " | " + obj.Country + " | " + str(
                                            obj.Latitude) + " | " + str(
                                            obj.Longitude))

                                        result = False

                                        if len(listLink) > 0:
                                            for i in range(len(listLink)):
                                                # print(str(obj.Title) + " " + str(listLink[z].Title))
                                                if (str(obj.Url) == str(listLink[i].Url) and str(obj.Latitude) == str(
                                                        listLink[i].Latitude) and str(obj.Longitude) == str(
                                                        listLink[i].Longitude)):
                                                    result = True
                                                    break

                                        if result == False:
                                            listLink.append(obj)
                                    except:
                                        continue
                                    break
                            else:
                                continue
                        else:
                            continue
                    except:
                        # listError.append(URL)
                        continue
                    break
            else:
                output = output['data']

                if len(output) > 0:
                    a = 0
                    for store in output:
                        a = a + 1
                        try:

                            obj = OBJ()
                            obj.Title = str(store['name'])
                            obj.Address = str(store['address'])

                            obj.City = str(store['city'])
                            obj.State = str(store['state'])
                            obj.Postcode = str(store['zipcode'])
                            obj.Country = "Australia"
                            obj.Latitude = str(store['geolocationLat'])
                            obj.Longitude = str(store['geolocationLong'])
                            obj.Url = str(store['urlPrefix'])
                            obj.PropertyId = str(store['propertyId'])

                            print(obj.Title + " | " + obj.City + " | " + obj.State + " | " + str(
                                obj.Postcode) + " | " + obj.Country + " | " + str(
                                obj.Latitude) + " | " + str(
                                obj.Longitude))

                            result = False

                            if len(listLink) > 0:
                                for i in range(len(listLink)):
                                    # print(str(obj.Title) + " " + str(listLink[z].Title))
                                    if (str(obj.Url) == str(listLink[i].Url) and str(obj.Latitude) == str(
                                            listLink[i].Latitude) and str(obj.Latitude) == str(
                                            listLink[i].Latitude) and str(obj.Longitude) == str(listLink[i].Longitude)):
                                        result = True
                                        break

                            if result == False:
                                listLink.append(obj)
                        except:
                            continue
                        break
                else:
                    continue

            break
    except:
        # listError.append(URL)
        continue


for z in range(len(listLink)):
    URL = 'https://d3ann4aa9tz2wb.cloudfront.net/api/v1/metadata/properties/#/listings'

    try:

        print(str(z) + " of " + str(len(listLink)) + " | " + URL.replace('#',listLink[z].PropertyId))
        res = requests.request("GET", URL.replace('#',listLink[z].PropertyId))

        if res.status_code == 200:
            output = json.loads(res.text)

            output = output['listings']

            if len(output) > 0:
                a = 0
                for store in output:
                    a = a + 1
                    try:
                        ID = str(store['id'])

                        obj = OBJ()
                        obj.Title = str(listLink[z].Title)
                        obj.Address = str(store['address'])
                        obj.FullAddress = str(listLink[z].Address)
                        obj.City = str(listLink[z].City)
                        obj.State = str(listLink[z].State)
                        obj.Postcode = str(listLink[z].Postcode)
                        obj.Country = "Australia"
                        obj.Latitude = str(listLink[z].Latitude)
                        obj.Longitude = str(listLink[z].Longitude)
                        obj.OfficeType = str(store['spaceTypeFormula'])
                        obj.Desk = str(store['noOfWS'])

                        print(str(obj.Desk) + " | " + str(obj.OfficeType) + " | " +  obj.Title + " | " + obj.City + " | " + obj.State + " | " + str(
                            obj.Postcode) + " | " + obj.Country + " | " + str(
                            obj.Latitude) + " | " + str(
                            obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                # print(str(obj.Title) + " " + str(listOBJ[z].Title))
                                if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.OfficeType) == str(
                                        listOBJ[i].OfficeType) and str(obj.Desk) == str(listOBJ[i].Desk) and str(
                                        obj.Address) == str(listOBJ[i].Address) and str(obj.City) == str(listOBJ[i].City) and str(
                                    obj.Latitude) == str(listOBJ[i].Latitude) and str(
                                    obj.Longitude) == str(listOBJ[i].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)
                    except:
                        continue
    except:
        continue

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(str(j))
    print(str(j) + " of " + str(len(listOBJ)) + " | " + str(listOBJ[z].Desk) + " | " + str(listOBJ[z].OfficeType) + " | " + " | " + listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(
        listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Desk)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].OfficeType)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=11).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents/OfficeHub_AUS_10062021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/OfficeHub_AUS_10062021.xlsx")
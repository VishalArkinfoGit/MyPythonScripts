from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime
import string

now = datetime.now()

my_path = "Documents/School_UAE_14052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""

listOBJ = []
listError = []

dt_string = now.strftime("%d%m%Y")

listLink = []


for z in range(1, 6):

    last = 1

    url = "https://whichschooladvisor.com/uae/school-search?school_phase=" + str(z) + "&page=1"
    res = requests.request("GET", url)

    if res.status_code == 200:
        soup = BeautifulSoup(res.content, "html.parser")

        try:
            list = soup.find('ul', class_='pagination').find_all("li")

            lastLi = list[len(list) - 1]

            output = str(lastLi.find("a")['href'])
            index1 = output.index('page=')
            last = int(output[index1 + len('page='):])
        except:
            last = 1

    for t in range(1, last+1):
        url = "https://whichschooladvisor.com/uae/school-search?school_phase=" + str(z) + "&page=" + str(t)
        print(url)

        try:

            res = requests.request("GET", url)

            if res.status_code == 200:
                soup = BeautifulSoup(res.content, "html.parser")

                try:
                    list = soup.find_all('section', class_='wsa-lm')

                    for li in list:
                        try:
                            obj = OBJ()

                            obj.Title = li.find('div', class_="wsa-lm-title").text

                            obj.Address = li.find('div', class_="wsa-lm-subtitle").text
                            obj.FullAddress = li.find('a', class_="wsa-nostyle-link")['href']

                            try:
                                res2 = requests.request("GET", obj.FullAddress+'/contact-information')

                                if res2.status_code == 200:
                                    soup2 = BeautifulSoup(res2.content, "html.parser")

                                    obj.Latitude = soup2.find('div', id='zena-googlemap')['data-lat']
                                    obj.Longitude = soup2.find('div', id='zena-googlemap')['data-lng']
                            except:
                                obj.Latitude = ''
                                obj.Longitude = ''


                            try:
                                listStat = soup.find("div", class_="wsa-lm-right-bottom").find_all("div",
                                                                                                   class_="wsa-lm-stat")

                                for a in listStat:
                                    try:
                                        if ("school_phase" in str(a.find("a")['href'])):
                                            output = str(a.find("a")['href'])
                                            index1 = output.index('=')
                                            obj.FullAddress = output[index1 + 1:]
                                            obj.FullAddress = obj.FullAddress + "|" + str(a.find("a").text)

                                    except:
                                        continue

                            except:
                                obj.Postcode = ""

                            print(obj.Title + " | " + obj.Address+" | "+ str(obj.Latitude) +" | "+ str(obj.Longitude))

                            result = False

                            if len(listOBJ) > 0:
                                for i in range(len(listOBJ)):
                                    if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                            listOBJ[i].Address)):
                                        result = True
                                        break

                            if result == False:

                                try:
                                    y = str(obj.Title) + ' ' + str(obj.Address)

                                    location = geolocator.geocode(
                                        y.translate(str.maketrans('', '', string.punctuation)))

                                    obj.Latitude = str(location.latitude)
                                    obj.Longitude = str(location.longitude)
                                except:
                                    obj.Latitude = ''
                                    obj.Longitude = ''

                                listOBJ.append(obj)

                        except:
                            continue

                except:
                    listError.append(url)
                    continue

            # break

        except:
            print("*******************************************")
            print("Location Not Found: " + url)
            listError.append(url)
            print("*******************************************")
            print("\n")
            continue


j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Latitude) + " | " + str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)
    sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].Type)

    wb_obj_w.save("Documents/School_UAE_14052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/School_UAE_14052021.xlsx")
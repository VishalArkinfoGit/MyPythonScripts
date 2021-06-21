from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
geolocator = Nominatim(user_agent="MyGeoCoder")

my_path = "Documents\ElectronicCity_Indonesia_25052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
listError = []
listLink = []
listPaginationLink = []
listFinalLink = []

url = 'https://eci.id/store/location'
res = requests.get(url)

if res.status_code == 200:
    soup = BeautifulSoup(res.content, "html.parser")
    output = soup.find(id="portfolio-list").find_all('li')

    for li in output:
        link = li.find('a')['href']

        listLink.append(link)



for link in listLink:
    print('https://eci.id' + link)
    res2 = requests.get('https://eci.id' + link)

    if res2.status_code == 200:
        soup = BeautifulSoup(res2.content, "html.parser")
        # print(soup.prettify())
        output = soup.find(id="loc_listv").find_all('div',class_="image-service-box")

        # for div in output:
        #     link2 = div.find('a', class_="pull-right")['href']
        #
        #     listFinalLink.append(link2)

        try:
            li = soup.find(class_="pagination").find('ul', id='yw0').find('li', class_='last')

            lastPageLink=li.find('a')['href']


            index1 = lastPageLink.index('page=')
            totalPage = lastPageLink[index1+len('page='):]

            for i in range(1, int(totalPage)+1):
                listPaginationLink.append(link + '?StoreLocation_page=' + str(i))

        except:
            listPaginationLink.append(link)

            continue
    else:
        continue

for link in listPaginationLink:
    print('https://eci.id' + link)
    res2 = requests.get('https://eci.id' + link)

    try:
        if res2.status_code == 200:
            soup2 = BeautifulSoup(res2.content, "html.parser")
            # print(soup.prettify())
            output = soup2.find(id="loc_listv").find_all('div', class_="image-service-box")

            for div in output:
                try:
                    obj = OBJ()

                    obj.Title = div.find('h4').text
                    obj.Address = str(div.find('p').text).replace("\t", "").replace("  ", "").strip().split('<a')[0]

                    obj.Address = str(obj.Address).replace('EC Connect', '').replace('021 -', '').replace('1500032','').replace('See Detail', '').strip()

                    print(obj.Title + " | " + obj.Address)

                    result = False

                    if len(listOBJ) > 0:
                        for z in range(len(listOBJ)):
                            if (str(obj.Title) == listOBJ[z].Title and str(obj.Address) == listOBJ[z].Address):
                                result = True
                                break

                    if result == False:
                        try:
                            obj.FullAddress = div.find('a', class_="pull-right")['href']

                            res3 = requests.get('https://eci.id' + obj.FullAddress)

                            if res3.status_code == 200:
                                soup3 = BeautifulSoup(res3.content, "html.parser")

                                index1 = str(soup3).index('google.maps.LatLng')
                                index1 = str(soup3).index('(', index1)
                                index2 = str(soup3).index(')', index1)
                                latlong = str(soup3)[index1 + 1:index2 - 1]

                                obj.Latitude = latlong.split(',')[0]
                                obj.Longitude = latlong.split(',')[1]

                        except:
                            obj.Latitude = ''
                            obj.Latitude = ''

                        obj.FullAddress =''
                        listOBJ.append(obj)
                except:
                    continue
        else:
            listError.append(link)
            continue
    except:
        listError.append(link)
        continue


j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Latitude) + " | " + str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\ElectronicCity_Indonesia_25052021.xlsx")

j = j + 10

if (len(listError) > 0):

    print('ERROR')

    for z in listError:
        # j = j + 1
        # sheet_obj_w.cell(row=j, column=1).value = str(z)
        # wb_obj_w.save("Documents\ElectronicCity_Indonesia_25052021.xlsx")
        print(str(z))
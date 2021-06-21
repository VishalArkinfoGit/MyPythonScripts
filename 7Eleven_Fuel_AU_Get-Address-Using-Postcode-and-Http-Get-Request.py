from math import trunc
from os import stat
import openpyxl
import requests 

import json
from types import SimpleNamespace

my_path = "C:/Python/Documents/Australian_All_Postcodes_11022021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row
print("Total Row: " + str(postcode_max_row))


my_path = "C:/Python/Documents/7_Eleven_Fuel_AU_05-04-2021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

# for i in range(2, 3170):
#     print(str(int(float(postcode_sheet_obj.cell(row = i, column = 2).value))))

j = 0

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
listUnique = []
obj = OBJ()
listOBJ.append(obj)

for i in range(2, 3169 + 1):
    # print(str(postcode_sheet_obj.cell(row = i, column = 2).value))

    try:
        URL = 'https://www.7eleven.com.au/storelocator-retail/mulesoft/stores?lat='+str(postcode_sheet_obj.cell(row = i, column = 5).value)+'&lng='+str(postcode_sheet_obj.cell(row = i, column = 4).value)+'&dist=10'
        #URL = 'https://www.7eleven.com.au/storelocator-retail/mulesoft/stores?lat=-37.8152065&lng=144.963937&dist=1000'

        print(URL)
       
        
        response = requests.get(URL).json()
        z = json.dumps(response)
        z = json.loads(z)
        #print(z)
        #print(response)
        # print(type(response))
        #print(response)
        

        #z = str(response).replace("\'", "\"").replace("None", "null").replace("True", "true").replace("False", "false")
        #print(z["stores"])

        #z = json.loads(z, object_hook=lambda d: SimpleNamespace(**d))
        
        #print(type(z.stores))

        listLocation = []

        listLocation = z["stores"]
        #print((listLocation))

        for y in range(len(listLocation)):
            x = listLocation[y]
            print(str(x["location"][0]))
            
            obj = OBJ()
            obj.Title = str(x["name"])
            obj.Address = str(x["address"]["address1"])
            obj.Suburb = str(x["address"]["suburb"])
            obj.State = str(x["address"]["state"])
            obj.Postcode = str(x["address"]["postcode"])
            obj.Country = str(x["region"]["countryId"])
            obj.Latitude = str(x["location"][0])
            obj.Longitude = str(x["location"][1])
        
            # print(obj.Address +" "+ str(obj.Latitude) +" "+ str(obj.Longitude))
            
            result = False

            if len(listOBJ) > 0:
                for z in range(len(listOBJ)):
                    if(str(obj.Latitude) == listOBJ[z].Latitude and str(obj.Longitude) == listOBJ[z].Longitude):
                        result = True
                        break
                    
            if result == False:
                listOBJ.append(obj)
                print("Success")

    except : 
        print("Exception")
        print(str(postcode_sheet_obj.cell(row = i, column = 5).value))


print(listOBJ)
for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" "+listOBJ[z].Address +" "+ str(listOBJ[z].Latitude) +" "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row = j, column = 5).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row = j, column = 6).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row = j, column = 7).value = str(str(listOBJ[z].Latitude))
    sheet_obj_w.cell(row = j, column = 8).value = str(str(listOBJ[z].Longitude))
    wb_obj_w.save("C:/Python/Documents/7_Eleven_Fuel_AU_05-04-2021.xlsx")
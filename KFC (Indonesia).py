
from bs4 import BeautifulSoup
import requests, json
import openpyxl

my_path = "C:/Python/Documents/KFC_Indonesia.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
obj = OBJ()
listOBJ.append(obj)

for i in range(1, 123+1):
    print(i)
    # URL = 'https://agents.helloworld.com.au/search-location/' + str(postcode_sheet_obj.cell(row=i, column=1).value)
    URL = 'https://kfcku.com/api/stores?page='+str(i)
    res = requests.get(URL)
    # res = requests.get(URL)
    # print(res.text)
    if res.status_code == 200:
            output = res.text
            print(output)

            if len(output['data']) > 0:
                for store in output['data']:

                    try:

                        obj = OBJ()
                        # print(store['ContactInfo']['FullAddress'])
                        obj.Title = str(store['name'])
                        obj.Address = str(store['address'])
                        obj.Latitude = str(store['long'])
                        obj.Longitude = str(store['lat'])

                        print(obj.Title + "|" + str(obj.Latitude) + " | " + str(
                            obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                # print(str(obj.Title) + " " + str(listOBJ[z].Title))
                                if (str(
                                    obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                                    listOBJ[i].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)
                    except:
                        continue

j=0
print(len(listOBJ))
for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" "+ str(listOBJ[z].Latitude) +" "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    # sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].Suburb)
    # sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].State)
    # sheet_obj_w.cell(row = j, column = 5).value = str(listOBJ[z].Country)
    # sheet_obj_w.cell(row = j, column = 6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row = j, column = 7).value = str(str(listOBJ[z].Latitude))
    sheet_obj_w.cell(row = j, column = 8).value = str(str(listOBJ[z].Longitude))
    # wb_obj_w.save("C:/Python/Documents/KFC_Indonesia.xlsx")

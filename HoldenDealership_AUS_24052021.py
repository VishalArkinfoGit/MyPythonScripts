from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json

my_path = "Documents\Australian_All_Suburb_13052021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row




my_path = "Documents\HoldenDealership_AUS_24052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []

headers = {
  'Cookie': 'ak_bmsc=B8D5C262437595C904E40415D86BD74C6862033DDE0A0000CFD8AB60EDA29357~plbxagI5hW8Eo2mPGxZ7jl8YWw1ZNFbFvs3AbSQRO5g2TGaFAUazmqQN3dXOuYHFE0Q2vgIZPn8ippDWBR2SWWBhnogaplizdw+u6K+Iy2cfjzSSgjurisOWvmLKIJx8IXStyT0KA7RDOwJtHfRFJerBd4Sf+ijDrczVj2ax9YUBl2SD7CDEHj5bmv4TeFqjyWkpzsOTAgLfyws3XB4s0PR2ihO7j9TOx9AYMsk90m6FY=; bm_sv=E1C5EB41BBDCCB4C622FDB7E6CFD81D2~QBbJV66xH1UR6XEA8yWFylvtgX9LRzR/euQuxegH8SW3omgk3PSDJWvHR/uC8UOM+VcbBJHc4+Pwm1oPJ+KYuuSVr0H6I+ZqcARl+83XI5MOJqBuVIOYVHOOCU1deU4lVCogDdg3RCAX/SZiTCxzlg=='
}

for i in range(1, postcode_max_row + 1):
    URL = 'https://oss.gm.com/api/v1/AU/dealerSearch/geoCoordinates?latitude='+\
          str(postcode_sheet_obj.cell(row = i, column = 4).value)+'&longitude='+\
          str(postcode_sheet_obj.cell(row = i, column = 5).value)+'&locale=en-AU'
    print(URL)
    try:

        res = requests.request("POST", URL, headers=headers)

        if res.status_code == 200:
            output = json.loads(res.text)

            output = output['dealerships']

            if len(output) > 0:
                for store in output:

                    try:

                        obj = OBJ()
                        obj.Title = str(store['name'])
                        obj.Address = str(store['address']['addressLine1'])

                        try:
                            if (store['address']['addressLine2'] != None):
                                obj.Address = obj.Address +", "+str(store['address']['addressLine2'])
                        except:
                            obj.Address = obj.Address

                        obj.City = str(store['address']['cityName'])
                        obj.State = str(store['address']['countrySubentity'])
                        obj.Postcode = str(store['address']['postalZone'])
                        obj.Country = "Australia"
                        obj.Latitude = str(store['latitude'])
                        obj.Longitude = str(store['longitude'])

                        print(obj.Title + " | " + obj.City + " | " + obj.State + " | " + str(
                            obj.Postcode) + " | " + obj.Country + " | " + str(obj.Latitude) + " | " + str(
                            obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                # print(str(obj.Title) + " " + str(listOBJ[z].Title))
                                if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                        listOBJ[i].Address) and str(obj.City) == str(listOBJ[i].City) and str(
                                    obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                                    listOBJ[i].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)
                    except:
                        continue
    except:
        # listError.append(URL)
        continue

    break

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\HoldenDealership_AUS_24052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\HoldenDealership_AUS_24052021.xlsx")
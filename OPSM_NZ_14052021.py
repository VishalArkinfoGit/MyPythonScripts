from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime

my_path = "Documents\OPSM_NZ_14052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active



class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []

listRegien = ['CAN','HKB','AUK','OTA','WKO','WGN','MBH','STL','NSN','MWT','NTL','BOP','TKI','WTC','GIS','TAS']

payload={}
headers = {
  'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
  'Accept-Encoding': 'gzip, deflate, br',
  'Accept-Language': 'en-US,en;q=0.9',
  'Cache-Control': 'max-age=0',
  'Connection': 'keep-alive',
  'Cookie': 'salutation_10152=null; date_10152=null; lastName_10152=null; firstName_10152=null; mobileno_10152=null; year_10152=null; month_10152=null; addressSuggest_10152=null; address2_10152=null; address1_10152=null; email_10152=null; medicareRef_10152=null; medicareno_10152=null; JSESSIONID=0000qmq3-NVSiLOOrpmoSwi1_Nw:1caatplae; REFERRER=https%3A%2F%2Fwww.google.com%2F; WC_SESSION_ESTABLISHED=true; WC_PERSISTENT=NkZ79UJOLo%2BctzpQ%2FNszDfz8Bu5ouer9QWZQ9ZpXm%2B8%3D%3B2021-05-14+16%3A30%3A27.884_1620973827883-55913_10152; WC_AUTHENTICATION_-1002=-1002%2CGWUh5ZtQLS%2F89OqlITP3utgDcmSp5FrgoJDUdeJqQms%3D; WC_ACTIVEPOINTER=-99%2C10152; WC_USERACTIVITY_-1002=-1002%2C10152%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2C1577704331%2CNejT1nXXHfVa8Uu7F93WeTcWbtZs58zvufUaqDIkDPVvP0w7ooqusLDSL9SUK0N9JXBIpJuk%2FBSft0Se7tFyEvTSFKW0mB2OPTYs%2BDn9Fd%2BkfyFqfUG07m2ge4G5zqA%2FqKuj9ql6GQcder6441eL%2BcXlQJChfs3xG7vw1pdNsfQEAFOJQTWr9gvS1xYI0qrMajok0%2BRFm12iFAsKPDcdxQWn2CT%2FQRkchQm4IIwJluaPQPfvN3EHNZxIdW%2BVsoeM; WC_GENERIC_ACTIVITYDATA=[747189598%3Atrue%3Afalse%3A0%3AoebCblzzJACAs9hF1Ghev7B5zAR4debJD0dCTLq07ZU%3D][com.ibm.commerce.context.ExternalCartContext|null][com.ibm.commerce.context.entitlement.EntitlementContext|4000000000000000005%264000000000000000005%26null%26-2000%26null%26null%26null][com.ibm.commerce.store.facade.server.context.StoreGeoCodeContext|null%26null%26null%26null%26null%26null][com.ibm.commerce.catalog.businesscontext.CatalogContext|12601%26null%26false%26false%26false][CTXSETNAME|Store][com.ibm.commerce.context.base.BaseContext|10152%26-1002%26-1002%26-1][com.ibm.commerce.context.audit.AuditContext|1620973827883-55913][com.ibm.commerce.context.experiment.ExperimentContext|null][com.ibm.commerce.giftcenter.context.GiftCenterContext|null%26null%26null][com.ibm.commerce.context.globalization.GlobalizationContext|-99%26NZD%26-99%26NZD]; mt.v=2.1355292483.1620973827761; bm_mi=9D79652AF2AC261CE21A872DB3F5B251~W7zECcqS/+cdalbtS41JSea3zBtRYqGtx6FTSs8/pVVV1hb7mMkt2LYHMZ+5NnM/KouPzN5ahqixS8XN0fhjqZDzFyP58NibeZe9w3ULmC8j9o9loUpDNWABoluAMGuubf96PVFPp5SGz5JxdRzQZ+n90b/TQZgiHNXY43RuQr+mC0vm6KjRNWbDr+2YpAocjytFM2ESeQRxK6HvAKjlj3C5maogGVip8AFF8qTZKqpt4ucPwVn+xaYYT5vGFdlf; _gcl_au=1.1.737746057.1620973835; ak_bmsc=E95E77FE8B5F2757371177A66A213FE25C7B8E6E0C59000004199E60E108D873~plCB7PlcztoiluvuDbf/pWLpQXU3D54i8aJcNJ1ZmzaQgAcPv2ZLRCvFejyx0PSNHvojwIIVgv15Jvk/x6nJjnHSR6TFt6PQRpbbNmcW7uNqolAkuGvTomtayRV+MLxQCTfuUnyw+QA7pz4MjQQaXjQIn8ec+PQ3QzY60yIzIsCwkePMt2DX47MhQeZoc3P4dA9KeSpYS4uAwM9hy2g1vzWcVQboROtDS5NT4DGNAtCWNce6siF+YZZctmcMB4ZInj; _ga=GA1.3.1280289508.1620973839; _gid=GA1.3.617483648.1620973839; tealium_data_session_timeStamp=1620973840203; _hjid=52a6cfcf-393c-4f19-8b8d-5827b5ba72fe; _hjTLDTest=1; _hjFirstSeen=1; AMCVS_125138B3527845350A490D4C%40AdobeOrg=1; _cs_mk=0.4979918283635023_1620973842038; s_ecid=MCMID%7C61474431837193781563078918033777661199; AMCV_125138B3527845350A490D4C%40AdobeOrg=-1303530583%7CMCIDTS%7C18762%7CMCMID%7C61474431837193781563078918033777661199%7CMCAAMLH-1621578641%7C6%7CMCAAMB-1621578641%7CRKhpRz8krg2tLO6pguXWp5olkAcUniQYPHaMWWgdJ3xzPWQmdj0y%7CMCOPTOUT-1620981041s%7CNONE%7CMCAID%7CNONE%7CvVersion%7C3.3.0; s_cc=true; __insp_wid=1727969683; __insp_nv=true; __insp_targlpt=; __insp_targlpu=aHR0cHM6Ly93d3cub3BzbS5jby5uei9maW5kLXN0b3JlL2F1ay9hdWNrbGFuZA%3D%3D; __insp_norec_sess=true; _fbp=fb.2.1620973857951.789966963; TS01f7d19d=011e1a41af89b2e90c8b142d839dbe94f46187fd8e35fa9aa1dbe4a503971c3cff6a7c5d57c12d24185b261f4c1e52e39d76c013b2591db8342e1ce8a0baf34ef387bbe9b5a986b3304133e643b4da7ac4643686a970e3a15f85bbb077f2e969ff280434aa8c206b6876866f4b0378bd75881160eaf3b6f2898cbc350be03293022f8ba25636f64dd5eb38071e8b0cdef17f3d26d9457bc019e8fa50946a44700091b191ef369b93caeeeb425313d00e78338e42b71927d0ca2a4659ccaee4eac40ac54eea; utag_main=v_id:01796991e7310067d842265dabd803072002e06a0086e$_sn:1$_se:6$_ss:0$_st:1620976018889$ses_id:1620973840178%3Bexp-session$_pn:6%3Bexp-session$vapi_domain:opsm.co.nz$dc_visit:1$dc_event:6%3Bexp-session$dc_region:eu-central-1%3Bexp-session; __insp_slim=1620974219416; _uetsid=e9938e00b47d11eb98a02d47f269e9bf; _uetvid=e993da60b47d11ebb76411da9fb0e48e; bm_sv=A1B375C5553EB38FD5D461B7249272F0~PCyuipZvnJY/1q7ksNtiGBrPPgII36u3QcBa4/vGQ3bSl2mUZGDeggeHkEQRt1u6vxlKK5RxYemy4+zJolmFIZqzvDlk3zTwILUc8uAUY7V/goHfDl7hmeyME6brbm/pDU0+P6IE1DoGaTKpUiO/zGlD5ehh+49Ur3tfhIMvFoo=; TS01f7d19d=011e1a41af44a0e50a8bd37ea77cfbeb7c8259465d40087c621e7c95403dd23fe34ae730b7a0d2a2184f4a52ce9c35cc79d6332a5b0306fc2648aa7a9952f3021577256e4fe0a67bc28c672bd7414a36cb32498b835cf1c3c4de4c7136cd4b7699b5b20a15c4f4f437a4bcab57c209c8a322915aed16798ecfad36cc08d8229d7f7777540fc5840b4f372c26dbb066282dcf23e22cde41f5f9733f4a807a2498c035c465d69f071315577a25386a641ff19b79dd16ffb4edd3f5f724f668a3a79f97a29df0; bm_sv=A1B375C5553EB38FD5D461B7249272F0~PCyuipZvnJY/1q7ksNtiGBrPPgII36u3QcBa4/vGQ3bSl2mUZGDeggeHkEQRt1u6vxlKK5RxYemy4+zJolmFIZqzvDlk3zTwILUc8uAUY7XI7TAjNKwsypEJ+Q4nPf2i/yjl0dvzimSzXZxZcqm4sBRMKfx/eg6VJtd5l+9qzAU=',
  'DNT': '1',
  'Host': 'www.opsm.co.nz',
  'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
  'sec-ch-ua-mobile': '?0',
  'Sec-Fetch-Dest': 'document',
  'Sec-Fetch-Mode': 'navigate',
  'Sec-Fetch-Site': 'none',
  'Sec-Fetch-User': '?1',
  'Upgrade-Insecure-Requests': '1',
  'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'
}

for reg in listRegien:
    url = "https://www.opsm.co.nz/find-store/" + reg

    res = requests.request("GET", url, headers=headers, data=payload)

    if res.status_code == 200:
        soup = BeautifulSoup(res.content, "html.parser")

        try:
            output = soup.find('div', class_='select-store--list').find_all('div', class_="select-store--item")

            for x in range(len(output)):
                obj = OBJ()

                obj.Title =str(output[x].find('input', id='description_'+str(x+1))['value']).strip()
                obj.Address =str(output[x].find('input', id='addressLine1_'+str(x+1))['value']).strip()
                obj.FullAddress =str(output[x].find('input', id='addressLine2_'+str(x+1))['value']).strip()
                obj.City =str(output[x].find('input', id='city_'+str(x+1))['value']).strip()
                obj.State =str(output[x].find('input', id='state_'+str(x+1))['value']).strip()
                obj.Postcode =str(output[x].find('input', id='postalCode_'+str(x+1))['value']).strip()
                obj.Latitude =str(output[x]['data-lat']).strip()
                obj.Longitude =str(output[x]['data-lng']).strip()

                print(obj.Title + " | " + obj.Address + " | " + obj.City + " | " + obj.State + " | " + str(obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                result = False

                if len(listOBJ) > 0:
                    for i in range(len(listOBJ)):
                        if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                listOBJ[i].Address)):
                            result = True
                            break

                if result == False:
                    listOBJ.append(obj)

        except:
            listError.append(url)
            continue

    # break



j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + listOBJ[z].City + " | " + listOBJ[z].State + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents/OPSM_NZ_14052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/OPSM_NZ_14052021.xlsx")
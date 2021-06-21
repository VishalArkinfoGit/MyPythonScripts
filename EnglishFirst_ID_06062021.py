from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime

my_path = "Documents\EnglishFirst_ID_06062021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active



class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []



listRegien = ['Bali', 'Balikpapan', 'Bandung', 'Banjarmasin', 'Bekasi', 'Bogor', 'Cikarang', 'Cilegon', 'Cimahi', 'Cirebon', 'Depok', 'Gresik', 'Jakarta', 'Jember', 'Kediri', 'Kupang', 'Lampung', 'Makassar', 'Malang', 'Manado', 'Mataram', 'Medan', 'Padang', 'Palembang', 'Pekalongan', 'Pekanbaru', 'Pontianak', 'Purwokerto', 'Samarinda', 'Semarang', 'Sidoarjo', 'Surabaya', 'Solo', 'Tangerang', 'Yogyakarta']

for reg in listRegien:
    url = "https://www.ef.co.id/englishfirst/kids/city/" + reg.lower() + "/"
    print(url)
    res = requests.request("GET", url)

    if res.status_code == 200:
        soup = BeautifulSoup(res.content, "html.parser")

        try:
            # output = str(soup.prettify()).replace('\n','').replace('\r','').replace('\t','').replace('  ','')

            output = soup.find('script',id='__NEXT_DATA__').contents[0]

            zz = str(output).replace("\\n", "").replace("\\t", "").replace("  ", "").strip()

            output = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

            locations = output.props.pageProps.pageContent.centers_data.centers

            for x in locations:
                try:
                    obj = OBJ()
                    obj.Title = str(x.DisplayName)
                    obj.Address = str(x.Address)
                    obj.FullAddress = str(x.Coordinate)
                    obj.City = str(x.LocalCityName)
                    obj.Suburb = str(x.District)

                    print(obj.Title + " | " + obj.Address + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                    result = False

                    if len(listOBJ) > 0:
                        for z in range(len(listOBJ)):
                            if (str(obj.Title) == listOBJ[z].Title and str(obj.Address) == listOBJ[
                                z].Address and str(obj.Latitude) == str(listOBJ[z].Latitude) and str(obj.Longitude) == str(listOBJ[
                                                                                    z].Longitude)):
                                result = True
                                break

                    if result == False:
                        listOBJ.append(obj)
                except:
                    continue
        except:
            listError.append(url)
            continue

    # break



j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Latitude) + " | " + str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].FullAddress)
    # sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    # sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    # sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    # sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents/EnglishFirst_ID_06062021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/EnglishFirst_ID_06062021.xlsx")
from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime
import string

now = datetime.now()

my_path = "Documents\EventCinemas_AUS_20052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""

listOBJ = []
listError = []

url = 'https://www.eventcinemas.com.au/Cinemas'

print(url)

try:
    res = requests.request("GET", url)

    if res.status_code == 200:
        soup = BeautifulSoup(res.content, "html.parser")

        try:
            output = soup.find_all('a',class_='cinema')

            for x in output:
                obj = OBJ()

                obj.Address = x['href']
                obj.State = x['data-state']
                obj.Latitude = x['data-long']
                obj.Longitude = x['data-lat']

                try:
                    res2 = requests.request("GET", 'https://www.eventcinemas.com.au'+obj.Address)
                    print('https://www.eventcinemas.com.au'+obj.Address)
                    if res2.status_code == 200:
                        soup2 = BeautifulSoup(res2.content, "html.parser")

                        index1 = str(soup2).index('{"address"')
                        index2 = str(soup2).index('</script>', index1)
                        xz = str(soup2)[index1:index2]

                        zz = str(xz).replace("\\n", "").replace("\\t", "").replace("  ", "").strip()

                        location = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

                        obj.Title = str(location.name)
                        obj.Address = str(location.address)
                        obj.Postcode = str(location.geo.postalCode)

                        print(obj.Title + " | " + obj.Address + " | " + obj.State + " | " + str(
                            obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(
                            obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                # print(str(obj.Title) + " " + str(listOBJ[z].Title))
                                if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                        listOBJ[i].Address) and str(obj.State) == str(listOBJ[i].State) and str(
                                    obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                                    listOBJ[i].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)
                except:
                    continue

        except:
            print("Error")
except:
    print("Error")

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + listOBJ[z].State + " | " + str(
        listOBJ[z].Latitude) + " | " + str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].Suburb)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)
    sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].Type)

    wb_obj_w.save("Documents\EventCinemas_AUS_20052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\EventCinemas_AUS_20052021.xlsx")
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
geolocator = Nominatim(user_agent="MyGeoCoder")
from googletrans import Translator
translator = Translator()
import string
from bs4 import BeautifulSoup
import requests
import openpyxl


my_path = "Documents\Reliance_IND_15062021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
  Title = ""
  Address = ""
  FullAddress = ""
  Suburb = ""
  State = ""
  City = ""
  Country = ""
  Postcode = ""
  Latitude = ""
  Longitude = ""
  Type = ""

listOBJ = []
listDistrict = [{"States":"Andaman And Nicobar (1)","Latitude":11.66702557,"Longitude":92.73598262},{"States":"Andhra Pradesh (20)","Latitude":14.7504291,"Longitude":78.57002559},{"States":"Arunachal Pradesh (1)","Latitude":27.10039878,"Longitude":93.61660071},{"States":"Assam (6)","Latitude":26.7499809,"Longitude":94.21666744},{"States":"Bihar (6)","Latitude":25.78541445,"Longitude":87.4799727},{"States":"Chandigarh (1)","Latitude":30.71999697,"Longitude":76.78000565},{"States":"Chhattisgarh (3)","Latitude":22.09042035,"Longitude":82.15998734},{"States":"Dadra And Nagar Haveli (12)","Latitude":20.26657819,"Longitude":73.0166178},{"States":"Delhi (2)","Latitude":28.6699929,"Longitude":77.23000403},{"States":"Goa (1)","Latitude":15.491997,"Longitude":73.81800065},{"States":"Haryana (10)","Latitude":28.45000633,"Longitude":77.01999101},{"States":"Himachal Pradesh (2)","Latitude":31.10002545,"Longitude":77.16659704},{"States":"Jammu And Kashmir (4)","Latitude":34.29995933,"Longitude":74.46665849},{"States":"Jharkhand (3)","Latitude":23.80039349,"Longitude":86.41998572},{"States":"Karnataka (16)","Latitude":12.57038129,"Longitude":76.91999711},{"States":"Kerala (5)","Latitude":8.900372741,"Longitude":76.56999263},{"States":"Lakshadweep (1)","Latitude":10.56257331,"Longitude":72.63686717},{"States":"Madhya Pradesh (8)","Latitude":21.30039105,"Longitude":76.13001949},{"States":"Maharashtra (20)","Latitude":19.25023195,"Longitude":73.16017493},{"States":"Manipur (1)","Latitude":24.79997072,"Longitude":93.95001705},{"States":"Meghalaya (1)","Latitude":25.57049217,"Longitude":91.8800142},{"States":"Mizoram (1)","Latitude":23.71039899,"Longitude":92.72001461},{"States":"Nagaland (1)","Latitude":25.6669979,"Longitude":94.11657019},{"States":"Orissa (6)","Latitude":19.82042971,"Longitude":85.90001746},{"States":"Puducherry (1)","Latitude":11.93499371,"Longitude":79.83000037},{"States":"Punjab (6)","Latitude":31.51997398,"Longitude":75.98000281},{"States":"Rajasthan (12)","Latitude":26.44999921,"Longitude":74.63998124},{"States":"Sikkim (1)","Latitude":27.3333303,"Longitude":88.6166475},{"States":"Tamil Nadu (19)","Latitude":12.92038576,"Longitude":79.15004187},{"States":"Tripura (1)","Latitude":23.83540428,"Longitude":91.27999914},{"States":"Uttar Pradesh (29)","Latitude":27.59998069,"Longitude":78.05000565},{"States":"Uttaranchal (1)","Latitude":30.32040895,"Longitude":78.05000565},{"States":"West Bengal (10)","Latitude":22.58039044,"Longitude":88.32994665}]
listNotFound = []


for x in listDistrict:
  url = 'https://storelocator.ril.com/getAllStores.aspx?flag=false&Searchformat=All&distance=1000&latitude=23.2410394&longitude=72.6476393'

  res = requests.request("GET", url)

  if res.status_code == 200:
    soup = BeautifulSoup(res.content, "html.parser")

    try:

      result = soup.text.split('$')

      for y in result:

        obj = OBJ()

        obj.Latitude = y.split('^')[1]
        obj.Longitude = y.split('^')[2]
        obj.Type = y.split('^')[3]
        obj.Title = y.split('^')[4]
        obj.FullAddress = y.split('^')[5]

        listOBJ.append(obj)
    except:
      print("*******************************************")
      print("Not Found: ")
      print("*******************************************")
      print("\n")

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(str(j) + " of " +str(len(listOBJ)) +" | "+listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].Type)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\Reliance_IND_15062021.xlsx")

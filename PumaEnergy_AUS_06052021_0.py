from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime
import string

now = datetime.now()


my_path = "Documents/Postcode.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Id = ""
    Suburb = ""
    Postcode = ""
    PostCodeName = ""
    Title = ""
    Address = ""
    FullAddress = ""
    # Suburb = ""
    State = ""
    City = ""
    Country = ""
    # Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""

listOBJ = []
listError = []

dt_string = now.strftime("%d%m%Y")

listAustraliaLocations = []
url = "https://www.pumaenergy.com.au/api/v1/suburbs/search/"

for z in range(10, -1, -1):
    payload = {'search': str(z),
               'stateID': '0',
               'limit': '99999'}

    print(url + str(z))

    res = requests.request("POST", url, headers=None, data=payload)

    if res.status_code == 200:
        output = json.loads(res.text)

        if (output['result']):

            for y in output['data']:
                obj = OBJ()

                obj.Title = y['suburb']['suburbID']
                obj.Suburb = y['suburb']['name']
                obj.Postcode = y['suburb']['postcode']
                obj.State = y['state']['name']


                try:
                    obj.FullAddress = str(obj.Suburb) + ' ' + str(obj.State) + ' ' + str(obj.Postcode) + ' Australia'

                    location = geolocator.geocode(obj.FullAddress.translate(str.maketrans('', '', string.punctuation)))
                    y = str(location.address).split(", ")

                    obj.Latitude = str(location.latitude)
                    obj.Longitude = str(location.longitude)
                except:
                    obj.Latitude = ""
                    obj.Longitude = ""

                print(obj.Suburb + " | " + obj.State + " | " + str(obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                result = False

                if len(listAustraliaLocations) > 0:
                    for z in range(len(listAustraliaLocations)):
                        if (str(obj.Title) == str(listAustraliaLocations[z].Title) and str(obj.Suburb) == str(listAustraliaLocations[z].Suburb) and str(obj.Postcode) == str(listAustraliaLocations[z].Postcode)):
                            result = True
                            break

                if result == False:
                    listAustraliaLocations.append(obj)


print(len(listAustraliaLocations))

# url = "https://www.pumaenergy.com.au/on-the-road/service-stations/stations/"
# for x in listAustraliaLocations:
#     payload = {'txtSuburbSearch': str(x.Suburb) +', '+str(x.State)+', '+str(x.Postcode),
#         'hdSearchType': 'serviceStations',
#         'hdSuburbID': str(x.Title),
#         'hdStateID': '0',
#         'btnSuburbSearch': 'Search Now'}
#
#     print(url + str(x.Suburb))
#
#     try:
#
#         res = requests.request("POST", url, headers=None, data=payload)
#
#         if res.status_code == 200:
#             soup = BeautifulSoup(res.content, "html.parser")
#
#             try:
#                 list = soup.find('div', class_="results").find("div", class_="listView").find_all('div',class_='serviceStation')
#
#                 for li in list:
#                     try:
#                         obj = OBJ()
#
#                         obj.Type = li['class']
#                         obj.FullAddress = li.find("a", class_="viewMore")['href']
#
#                         # print(obj.FullAddress)
#
#                         try:
#                             res = requests.get('https://www.pumaenergy.com.au/' + obj.FullAddress)
#                             soup = BeautifulSoup(res.content, "html.parser")
#
#                             spans = soup.find("div", class_="singleStation").find_all('span')
#
#                             for span in spans:
#                                 try:
#                                     if (span['itemprop'] == 'name'):
#                                         obj.Title = span.text
#                                     elif (span['itemprop'] == 'streetAddress'):
#                                         obj.Address = span.text
#                                     elif (span['itemprop'] == 'addressLocality'):
#                                         obj.Suburb = span.text
#                                     elif (span['itemprop'] == 'addressRegion'):
#                                         obj.State = span.text
#                                     elif (span['itemprop'] == 'postalCode'):
#                                         obj.Postcode = span.text
#                                 except:
#                                     obj.FullAddress = obj.FullAddress
#
#                             obj.FullAddress = ""
#
#                             try:
#                                 output = str(soup.prettify())
#
#                                 index1 = output.find('google.maps.LatLng')
#                                 index1 = output.find('(', index1)
#                                 index2 = output.find(')', index1)
#                                 output = output[index1+1: index2]
#
#                                 obj.Latitude = str(output.split(',')[0])
#                                 obj.Longitude = str(output.split(',')[1])
#                             except:
#                                 obj.FullAddress = ''
#                                 obj.Latitude = ''
#                                 obj.Longitude = ''
#                         except:
#
#                             obj.Title = li.find("div", class_="details").find('h2').find('a').text
#                             index1 = obj.Title.find('<span>')
#                             obj.Title = output[: index1 + 1]
#
#                             spans = soup.find("div", class_="address").find_all('span')
#
#                             for span in spans:
#                                 try:
#                                     if (span['itemprop'] == 'streetAddress'):
#                                         obj.Address = span.text
#                                     elif (span['itemprop'] == 'addressLocality'):
#                                         obj.Suburb = span.text
#                                     elif (span['itemprop'] == 'addressRegion'):
#                                         obj.State = span.text
#                                     elif (span['itemprop'] == 'postalCode'):
#                                         obj.Postcode = span.text
#
#                                 except:
#                                     obj.FullAddress = obj.FullAddress
#
#                             obj.FullAddress = ''
#
#
#                         print(obj.Title + " | " + obj.Address + " | " + str(obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))
#
#                         result = False
#
#                         if len(listOBJ) > 0:
#                             for i in range(len(listOBJ)):
#                                 if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(listOBJ[i].Address) and str(obj.Postcode) == str(listOBJ[i].Postcode) and str(obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(listOBJ[i].Longitude)):
#                                     result = True
#                                     break
#
#                         if result == False:
#                             listOBJ.append(obj)
#
#                     except:
#                         continue
#
#             except:
#                 listError.append(str(x.Suburb) +', '+str(x.State)+', '+str(x.Postcode))
#                 continue
#         # break
#
#     except:
#         print("*******************************************")
#         print("Location Not Found: " + str(x.Suburb) +', '+str(x.State)+', '+str(x.Postcode))
#         listError.append(str(x.Suburb) +', '+str(x.State)+', '+str(x.Postcode))
#         print("*******************************************")
#         print("\n")
#         continue
#
#
# j = 0

# for z in range(len(listOBJ)):
#     j = j + 1
#     print(listOBJ[z].Title +" | "+listOBJ[z].Address +" | "+ str(listOBJ[z].Postcode) +" | "+ str(listOBJ[z].Latitude) +" | "+ str(listOBJ[z].Longitude))
#     sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
#     sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
#     sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].City)
#     sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].State)
#     sheet_obj_w.cell(row = j, column = 5).value = str(listOBJ[z].Suburb)
#     sheet_obj_w.cell(row = j, column = 6).value = str(listOBJ[z].Postcode)
#     sheet_obj_w.cell(row = j, column = 7).value = str(listOBJ[z].Country)
#     sheet_obj_w.cell(row = j, column = 8).value = str(listOBJ[z].Latitude)
#     sheet_obj_w.cell(row = j, column = 9).value = str(listOBJ[z].Longitude)
#
#     wb_obj_w.save("Documents/Postcode.xlsx")
#
# j = j + 10



j = 0
if (len(listAustraliaLocations) > 0):

    for z in listAustraliaLocations:
        print(z.Suburb + " | " + z.State + " | " + str(z.Postcode) + " | " + str(z.Latitude) + " | " + str(z.Longitude))

        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z.Suburb)
        sheet_obj_w.cell(row=j, column=2).value = str(z.State)
        sheet_obj_w.cell(row=j, column=3).value = str(z.Postcode)
        sheet_obj_w.cell(row=j, column=4).value = str(z.Latitude)
        sheet_obj_w.cell(row=j, column=5).value = str(z.Longitude)
        sheet_obj_w.cell(row=j, column=6).value = str(z.Title)
        wb_obj_w.save("Documents/Postcode.xlsx")
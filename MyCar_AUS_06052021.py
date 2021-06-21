from bs4 import BeautifulSoup
import requests
import openpyxl
import requests 

import json
from types import SimpleNamespace


listStateUrl = ['https://www.mycar.com.au/stores/act','https://www.mycar.com.au/stores/nsw','https://www.mycar.com.au/stores/qld','https://www.mycar.com.au/stores/wa','https://www.mycar.com.au/stores/vic','https://www.mycar.com.au/stores/nt','https://www.mycar.com.au/stores/sa','https://www.mycar.com.au/stores/tas']

my_path = "Documents/CIBO_AU_06052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""

listOBJ = []
listError = []

for url in listStateUrl:

    try:
        print(url)

        res = requests.get(url)

        soup = BeautifulSoup(res.content, "html.parser")

        try:
            ul = soup.find("ul",id="storesGroup")

            for li in ul:
                try:
                    obj = OBJ()

                    obj.Title = str(li.find("span", class_="store-text").text).strip()
                    obj.Type = str(li.find("span", class_="store-details").text).strip()
                    obj.FullAddress = li.find("div", class_="btn-container").find('a')['href']

                    obj.Latitude = str(li['data-store-lat'])
                    obj.Longitude = str(li['data-store-lng'])
                    obj.State = str(li['data-store-state'])

                    obj.Country = "Australia"

                    print(obj.FullAddress)

                    res = requests.get(obj.FullAddress)

                    soup = BeautifulSoup(res.content, "html.parser")

                    try:
                        obj.Address = str(soup.find("div", class_="store-contact-info").find("span", class_="selectedstore").find("p").text).strip()
                        obj.Address = " ".join(obj.Address.split())
                    except:
                        listError.append(obj.FullAddress)
                        obj.Address = ''


                    print(obj.Title + " | " + obj.Address + " | " + obj.Country + " | " + str(
                        obj.Latitude) + " | " + str(obj.Longitude))

                    result = False

                    if len(listOBJ) > 0:
                        for i in range(len(listOBJ)):
                            if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(listOBJ[i].Address) and str(obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(listOBJ[i].Longitude)):
                                result = True
                                break

                    if result == False:
                        listOBJ.append(obj)

                except:
                    continue

        except:
            print("Exception: " + url)
            listError.append(url)
            continue

    except :
        print("Exception: " + url)
        listError.append(url)
        continue


j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" | "+listOBJ[z].Address +" | "+ str(listOBJ[z].Latitude) +" | "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row = j, column = 5).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row = j, column = 6).value = str(listOBJ[z].Longitude)
    sheet_obj_w.cell(row = j, column = 7).value = str(listOBJ[z].Type)
    wb_obj_w.save("Documents/CIBO_AU_06052021.xlsx")

j = j + 10

if (len(listError) > 0):
    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/CIBO_AU_06052021.xlsx")
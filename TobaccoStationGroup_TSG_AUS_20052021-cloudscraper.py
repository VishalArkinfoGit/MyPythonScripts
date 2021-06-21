from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json
import cloudscraper

scraper = cloudscraper.create_scraper()

my_path = "Documents\Australian_All_Suburb_QLD_13052021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row
print("Total Row: " + str(postcode_max_row))




my_path = "Documents\TobaccoStationGroup_TSG_AUS_20052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []




for i in range(1, postcode_max_row + 1):

    headers = {'accept': '*/*',
               'cookie': '_ga=GA1.2.554171629.1621525785; _gid=GA1.2.1958975876.1621525785; _gat_gtag_UA_24783719_1=1',
               'dnt': '1',
               'referer': 'https://tobaccosg.com/store-locator/',
               'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
               'sec-ch-ua-mobile': '?0',
               'sec-fetch-dest': 'empty',
               'sec-fetch-mode': 'cors',
               'sec-fetch-site': 'same-origin',
               'user-agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36',
               'x-requested-with': 'XMLHttpRequest'}

    URL = 'https://tobaccosg.com/wp-admin/admin-ajax.php?action=store_search&lat='+str(postcode_sheet_obj.cell(row = i, column = 4).value)+'&lng='+str(postcode_sheet_obj.cell(row = i, column = 5).value)+'&max_results=250&search_radius=500&autoload=1'
    print(URL)
    # try:
    if(len(URL) > 0):
        res = scraper.request("GET", URL, headers=headers)

        if res.status_code == 200:
            output = json.loads(res.text)

            if len(output) > 0:
                for store in output:

                    try:

                        obj = OBJ()
                        obj.Title = str(store['store'])
                        obj.Address = str(store['address'])
                        obj.FullAddress = str(store['address2'])
                        obj.City = str(store['city'])
                        obj.State = str(store['state'])
                        obj.Postcode = str(store['zip'])
                        obj.Country = str(store['country'])
                        obj.Latitude = str(store['lat'])
                        obj.Longitude = str(store['lng'])

                        print(obj.Title + " | " + obj.City + " | " + obj.State + " | " + str(
                            obj.Postcode) + " | " + obj.Country + " | " + str(obj.Latitude) + " | " + str(
                            obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                # print(str(obj.Title) + " " + str(listOBJ[z].Title))
                                if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                        listOBJ[i].Address) and str(obj.City) == str(listOBJ[i].City) and str(
                                    obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                                    listOBJ[i].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)
                    except:
                        continue
    # except:
    #     listError.append(URL)
    #     continue

    # break

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\TobaccoStationGroup_TSG_AUS_20052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\TobaccoStationGroup_TSG_AUS_20052021.xlsx")
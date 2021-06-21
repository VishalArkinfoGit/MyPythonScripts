from bs4 import BeautifulSoup
import requests
import openpyxl
import requests 

import json
from types import SimpleNamespace


listStateUrl = ['https://www.ciboespresso.com.au/api/theme/store/?state=south+australia','https://www.ciboespresso.com.au/api/theme/store/?state=new+south+wales','https://www.ciboespresso.com.au/api/theme/store/?state=queensland','https://www.ciboespresso.com.au/api/theme/store/?state=victoria','https://www.ciboespresso.com.au/api/theme/store/?state=western+australia']

my_path = "Documents/CIBO_AU_06052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
listError = []

for url in listStateUrl:

    try:
        print(url)

        response = requests.get(url).json()

        listLocation = response["posts"]

        for y in range(len(listLocation)):
            x = listLocation[y]

            obj = OBJ()
            obj.Title = str(x["name"])
            obj.Address = str(x["fields"]['address'])

            try:
                obj.FullAddress = str(x["fields"]["map"]['address'])
                obj.Latitude = str(x["fields"]["map"]["lat"])
                obj.Longitude = str(x["fields"]["map"]["lng"])
            except:
                obj.FullAddress = ''
                obj.Latitude = ''
                obj.Longitude = ''

            obj.Country = "Australia"


            print(obj.Title + " | " + obj.Address + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

            result = False

            if len(listOBJ) > 0:
                for i in range(len(listOBJ)):
                    if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.FullAddress) == str(
                            listOBJ[i].FullAddress) and str(obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(
                            listOBJ[i].Longitude)):
                        result = True
                        break

            if result == False:
                listOBJ.append(obj)

    except : 
        print("Exception: " + url)
        listError.append(url)
        continue


j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" "+listOBJ[z].Address +" "+ str(listOBJ[z].Latitude) +" "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row = j, column = 5).value = str(str(listOBJ[z].Latitude))
    sheet_obj_w.cell(row = j, column = 6).value = str(str(listOBJ[z].Longitude))
    wb_obj_w.save("Documents/CIBO_AU_06052021.xlsx")

j = j + 10

if (len(listError) > 0):
    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/CIBO_AU_06052021.xlsx")
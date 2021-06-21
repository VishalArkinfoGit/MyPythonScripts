from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json
from types import SimpleNamespace

my_path = "Documents\YellowPages_AUS_NSW_24052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

# for i in range(2, 3170):
#     print(str(int(float(postcode_sheet_obj.cell(row = i, column = 2).value))))


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""
    Type = ""


listOBJ = []
listError = []

headers = {
  'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
  'accept-encoding': 'gzip, deflate, br',
  'accept-language': 'en-US,en;q=0.9',
  'cache-control': 'max-age=0',
  'cookie': '_vwo_uuid_v2=D308924CE0C2A9CC17A9456C4557D61D5|8dbaa0ff82209af2ef672e4a6ba8326a; _vwo_uuid=D308924CE0C2A9CC17A9456C4557D61D5; s_ecid=MCMID|64569858247329101162770077466101935047; _wingify_pc_uuid=c057eecea3a1460da071b3266c2a3885; wingify_donot_track_actions=0; BVBRANDID=e6daa803-07c9-4a8f-8f76-55b101625a2d; _vis_opt_exp_217_combi=1; locationClue=Victoria; _vwo_ds=3:a_0,t_0:-1$1621612125:94.4529111:::3_0,2_0:0; RT="z=1&dm=www.yellowpages.com.au&si=14f886ca-6ab9-4832-8463-a444bbd3ba43&ss=kp5va2c3&sl=0&tt=0"; yellow-guid=c48be539-5dd7-445d-b0ae-d651b6716eee; ak_bmsc=18719201971131225AE05592064ADE7717C857694B640000C9A8AE609950D506~pluy/feYFH7bJQ1rfZADR9QrQasnrxoOeTZ/esmDKeqphIutYsVnNw+vhT8n+gvgkSbPQlCAGpUTczgsk0SdiS1bR0sAUNE5G1BtGEBTcvX7VPI+Ee0y3jAUR3DX21jexnU6NssytY6oKoicR10fHdlMq0F3WkhiDL/SoO27nXm0ApBJ+Neauaw4clLm6c/KK50vDo3ig33u2FztZ3y6pk2Srp6wqV0Rq1D3foI7ueaCpydkuIAn4f/KmNGAwd8w1Y; _vis_opt_s=8|; _vis_opt_test_cookie=1; _vwo_sn=447087:1; bm_sv=3729877ACA0375EDC2B405838D2E5025~0yRxgMR/hvEG+JzuyRdITQtl2k4YuFMq1L5QbkqsWnlQFXNXCjcJ+c/tPvYSoQpGLGvSCv26mnBZFKANfU4fuQ+pXZWZfajaUEKSFMoSjnkQFOL+eyH+aWRx9mUQl4a+Us9LR05f9Jb47lqYLRvvJnl+n4XMlRmpunsGIduEBMY=; AMCVS_8412403D53AC3D7E0A490D4C@AdobeOrg=1; AMCV_8412403D53AC3D7E0A490D4C@AdobeOrg=-1124106680|MCIDTS|18774|MCMID|64569858247329101162770077466101935047|MCAAMLH-1622664015|6|MCAAMB-1622664015|RKhpRz8krg2tLO6pguXWp5olkAcUniQYPHaMWWgdJ3xzPWQmdj0y|MCOPTOUT-1622066415s|NONE|MCAID|NONE|vVersion|5.2.0; s_cc=true',
  'dnt': '1',
  'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
  'sec-ch-ua-mobile': '?0',
  'sec-fetch-dest': 'document',
  'sec-fetch-mode': 'navigate',
  'sec-fetch-site': 'none',
  'sec-fetch-user': '?1',
  'upgrade-insecure-requests': '1',
  'user-agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'
}



def recursive(list, url):
    print(url)
    res = requests.request("GET", url, headers=headers)

    if res.status_code == 200:
        # soup = BeautifulSoup(res.content, "html.parser")

        try:
            output = str(res.text).replace("\\n", "").replace("\n", "").replace("\\t", "").replace("\t", "").replace(
                "  ", "")

            index1 = output.index('__INITIAL_STATE__')
            index1 = output.index('{', index1)
            index2 = output.index('</script>', index1)
            output = output[index1:index2 - 1]
            zz = str(output).replace("\\n", "").replace("\\t", "").replace("  ", "").strip()

            locations = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

            try:
                locations = locations.model.inAreaResultViews

                for x in locations:
                    try:
                        obj = OBJ()
                        obj.Title = str(x.name)
                        obj.Address = str(x.addressView.addressLine)
                        obj.FullAddress = str(x.addressView.asContactCardFormat)
                        obj.Suburb = str(x.addressView.suburb)
                        obj.State = str(x.addressView.state)
                        obj.Postcode = str(x.addressView.postCode)
                        obj.Country = "Australia"
                        obj.Latitude = str(x.addressView.latitude)
                        obj.Longitude = str(x.addressView.longitude)
                        obj.Type = str(x.category.name)

                        print(obj.Title + " | " + obj.Address + " | " + obj.Suburb + " | " + str(
                            obj.Postcode) + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                        result = False

                        if len(list) > 0:
                            for z in range(len(list)):
                                if (str(obj.Title) == list[z].Title and str(obj.Address) == list[
                                    z].Address and str(
                                    obj.Postcode) == str(list[z].Postcode) and str(obj.Latitude) == str(
                                    list[z].Latitude) and str(obj.Longitude) == str(list[
                                                                                           z].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            list.append(obj)
                    except:
                        continue
            except:
                print('Error')

        except:
            print(url)

    return list

listPostCode = [1001,1208,1216,1221,1231,1300,1335,1340,1355,1360,1401,1420,1430,1440,1450,1460,1466,1470,1475,1480,1490,1493,1495,1499,1515,1565,1570,1582,1590,1595,1630,1639,1655,2103,1675,1680,1685,1700,1710,1715,1730,1740,1750,1755,1765,1790,1800,1805,1811,1816,1830,1835,1848,1860,1871,1875,1885,1890,1891,2000,2002,2006,2007,2008,2009,2010,2011,2015,2016,2017,2018,2019,2020,2021,2022,2023,2024,2025,2026,2027,2029,2030,2031,2032,2034,2035,2036,2037,2038,2039,2040,2041,2042,2043,2044,2045,2046,2048,2049,2050,2055,2057,2060,2061,2062,2063,2065,2066,2067,2068,2069,2070,2071,2072,2073,2074,2075,2076,2077,2079,2080,2081,2082,2083,2084,2085,2086,2087,2088,2089,2090,2092,2093,2094,2095,2096,2097,2099,2100,2101,2102,2104,2105,2106,2107,2108,2109,2110,2111,2112,2113,2114,2116,2117,2118,2119,2120,2121,2122,2125,2126,2127,2129,2130,2132,2133,2134,2136,2137,2138,2139,2140,2141,2142,2143,2145,2146,2147,2148,2150,2151,2152,2153,2155,2156,2157,2158,2159,2160,2161,2162,2163,2164,2165,2166,2167,2168,2170,2171,2172,2173,2175,2176,2177,2178,2179,2190,2191,2192,2193,2194,2195,2196,2197,2198,2199,2200,2203,2204,2205,2206,2207,2209,2210,2211,2212,2213,2216,2217,2218,2219,2220,2221,2222,2223,2224,2225,2226,2227,2228,2229,2230,2231,2232,2233,2234,2250,2251,2256,2257,2258,2259,2260,2261,2262,2263,2264,2265,2267,2278,2280,2281,2282,2283,2284,2285,2286,2287,2289,2290,2291,2292,2293,2294,2295,2296,2297,2298,2299,2300,2302,2303,2304,2305,2306,2307,2308,2309,2311,2312,2314,2315,2316,2317,2318,2319,2320,2321,2322,2323,2324,2325,2326,2327,2328,2329,2330,2331,2333,2334,2335,2336,2337,2338,2339,2340,2341,2342,2343,2344,2345,2346,2347,2350,2351,2352,2353,2354,2355,2356,2357,2358,2359,2360,2361,2365,2369,2370,2371,2372,2379,2380,2381,2382,2386,2387,2388,2390,2395,2396,2397,2398,2399,2400,2401,2402,2403,2404,2405,2406,2408,2409,2410,2411,2415,2420,2421,2422,2423,2424,2425,2426,2427,2428,2429,2430,2431,2439,2440,2441,2443,2444,2445,2446,2447,2448,2449,2450,2452,2453,2454,2455,2456,2460,2462,2463,2464,2465,2466,2469,2470,2471,2472,2473,2474,2475,2476,2477,2478,2479,2480,2481,2482,2483,2484,2485,2486,2487,2488,2489,2490,2500,2502,2505,2506,2508,2515,2516,2517,2518,2519,2522,2525,2526,2527,2528,2529,2530,2533,2534,2535,2536,2537,2538,2539,2540,2541,2545,2546,2548,2549,2550,2551,2555,2556,2557,2558,2559,2560,2563,2564,2565,2566,2567,2568,2569,2570,2571,2572,2573,2574,2575,2576,2577,2578,2579,2580,2581,2582,2583,2584,2585,2586,2587,2588,2590,2594,2600,2601,2602,2603,2604,2605,2606,2607,2608,2609,2611,2612,2614,2615,2616,2617,2618,2619,2620,2621,2622,2623,2624,2625,2626,2627,2628,2629,2630,2631,2632,2633,2640,2641,2642,2643,2644,2645,2646,2647,2648,2649,2650,2651,2652,2653,2655,2656,2658,2659,2660,2661,2663,2665,2666,2668,2669,2671,2672,2675,2678,2680,2681,2700,2701,2702,2703,2705,2706,2707,2710,2711,2712,2713,2714,2715,2716,2717,2720,2721,2722,2725,2726,2727,2729,2730,2731,2732,2733,2734,2735,2736,2737,2738,2739,2745,2747,2748,2749,2750,2752,2753,2754,2755,2756,2757,2758,2759,2760,2761,2762,2763,2765,2766,2767,2768,2769,2770,2773,2774,2775,2776,2777,2778,2779,2780,2782,2783,2784,2785,2786,2787,2790,2791,2792,2793,2794,2795,2797,2798,2799,2800,2803,2804,2805,2806,2807,2808,2809,2810,2820,2821,2823,2824,2825,2827,2828,2829,2830,2831,2832,2833,2834,2835,2836,2839,2840,2842,2843,2844,2845,2846,2847,2848,2849,2850,2852,2864,2865,2866,2867,2868,2869,2870,2871,2873,2874,2875,2876,2877,2878,2879,2880,2898,2899]

j = 0

for code in listPostCode:
    if code == 1300:
        break

    URL = 'https://www.yellowpages.com.au/search/listings?clue=Vets+%26+Veterinary&locationClue='+str(code)+'&selectedViewMode=list'
    print(URL)
    try:
        res = requests.request("GET", URL, headers=headers)

        if res.status_code == 200:
            # soup = BeautifulSoup(res.content, "html.parser")

            try:
                output = str(res.text).replace("\\n", "").replace("\n", "").replace("\\t", "").replace("\t",
                                                                                                       "").replace("  ",
                                                                                                                   "")

                index1 = output.index('__INITIAL_STATE__')
                index1 = output.index('{', index1)
                index2 = output.index('</script>', index1)
                output = output[index1:index2 - 1]
                zz = str(output).replace("\\n", "").replace("\\t", "").replace("  ", "").strip()

                locations = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

                try:

                    totalResults = locations.model.pagination.totalResults

                    searchResultsPerPage = locations.model.pagination.searchResultsPerPage

                    totalPages = int(totalResults / searchResultsPerPage) + 1

                    for no in range(1, totalPages + 1):

                        list = recursive(listOBJ, URL + '&pageNumber=' + str(no))

                        listOBJ.extend(list)

                        try:
                            for z in range(len(list)):
                                j = j + 1
                                sheet_obj_w.cell(row=j, column=1).value = str(list[z].Title)
                                sheet_obj_w.cell(row=j, column=2).value = str(list[z].Address)
                                sheet_obj_w.cell(row=j, column=3).value = str(list[z].FullAddress)
                                sheet_obj_w.cell(row=j, column=4).value = str(list[z].Suburb)
                                sheet_obj_w.cell(row=j, column=5).value = str(list[z].State)
                                sheet_obj_w.cell(row=j, column=6).value = str(list[z].Postcode)
                                sheet_obj_w.cell(row=j, column=7).value = str(list[z].Country)
                                sheet_obj_w.cell(row=j, column=8).value = str(list[z].Latitude)
                                sheet_obj_w.cell(row=j, column=9).value = str(list[z].Longitude)
                                sheet_obj_w.cell(row=j, column=10).value = str(list[z].Type)

                                wb_obj_w.save("Documents\YellowPages_AUS_NSW_24052021.xlsx")
                        except:
                            wb_obj_w.save("Documents\YellowPages_AUS_NSW_24052021.xlsx")
                            continue

                except:
                    wb_obj_w.save("Documents\YellowPages_AUS_NSW_24052021.xlsx")
                    continue

            except:
                listError.append(URL)
                wb_obj_w.save("Documents\YellowPages_AUS_NSW_24052021.xlsx")
                continue

            # break

    except:
        wb_obj_w.save("Documents\YellowPages_AUS_NSW_24052021.xlsx")
        continue
# j = 0

# for z in range(len(listOBJ)):
#     j = j + 1
#     print(listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Latitude) + " | " + str(
#         listOBJ[z].Longitude))
#     sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
#     sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
#     sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
#     sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].Suburb)
#     sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
#     sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
#     sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
#     sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].Latitude)
#     sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].Longitude)
#     sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].Type)
#
#     wb_obj_w.save("Documents\YellowPages_AUS_NSW_24052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\YellowPages_AUS_NSW_24052021.xlsx")
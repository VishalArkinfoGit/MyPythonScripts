from bs4 import BeautifulSoup
import requests, json
import openpyxl
from types import SimpleNamespace
import ssl
import certifi
import geopy.geocoders
ctx = ssl.create_default_context(cafile=certifi.where())
geopy.geocoders.options.default_ssl_context = ctx
from geopy.geocoders import Nominatim
from openpyxl.descriptors.base import Integer
geolocator = Nominatim(user_agent="MyGeoCoder")
import xlsxwriter
from datetime import datetime

my_path = "Documents/BLCW_AUS_18062021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active



class OBJ:
    Postcode = ""
    Population = ""
    NDISSpendByServiceArea_Current_Key = ""
    NDISSpendByServiceArea_Current_Low = ""
    NDISSpendByServiceArea_Current_High = ""
    NDISSpendByServiceArea_Current_TotalLow = ""
    NDISSpendByServiceArea_Current_TotalHigh = ""
    NDISParticipantByType_Current_Key = ""
    NDISParticipantByType_Current_Low = ""
    NDISParticipantByType_Current_High = ""
    NDISParticipantByType_Current_TotalLow = ""
    NDISParticipantByType_Current_TotalHigh = ""
    NDISSpendByServiceArea_Forecast_Key = ""
    NDISSpendByServiceArea_Forecast_Low = ""
    NDISSpendByServiceArea_Forecast_High = ""
    NDISSpendByServiceArea_Forecast_TotalLow = ""
    NDISSpendByServiceArea_Forecast_TotalHigh = ""
    NDISParticipantByType_Forecast_Key = ""
    NDISParticipantByType_Forecast_Low = ""
    NDISParticipantByType_Forecast_High = ""
    NDISParticipantByType_Forecast_TotalLow = ""
    NDISParticipantByType_Forecast_TotalHigh = ""
    NDISTotalWorkforce_Forecast_Key = ""
    NDISTotalWorkforce_Forecast_Low = ""
    NDISTotalWorkforce_Forecast_High = ""
    NDISTotalWorkforce_Forecast_TotalLow = ""
    NDISTotalWorkforce_Forecast_TotalHigh = ""

class OBJ1:
    Postcode = ""
    Population = ""

class OBJ2:
    NDISSpendByServiceArea_Current_Key = ""
    NDISSpendByServiceArea_Current_Low = ""
    NDISSpendByServiceArea_Current_High = ""
    NDISSpendByServiceArea_Current_TotalLow = ""
    NDISSpendByServiceArea_Current_TotalHigh = ""

class OBJ3:
    NDISParticipantByType_Current_Key = ""
    NDISParticipantByType_Current_Low = ""
    NDISParticipantByType_Current_High = ""
    NDISParticipantByType_Current_TotalLow = ""
    NDISParticipantByType_Current_TotalHigh = ""

class OBJ4:
    NDISSpendByServiceArea_Forecast_Key = ""
    NDISSpendByServiceArea_Forecast_Low = ""
    NDISSpendByServiceArea_Forecast_High = ""
    NDISSpendByServiceArea_Forecast_TotalLow = ""
    NDISSpendByServiceArea_Forecast_TotalHigh = ""

class OBJ5:
    NDISParticipantByType_Forecast_Key = ""
    NDISParticipantByType_Forecast_Low = ""
    NDISParticipantByType_Forecast_High = ""
    NDISParticipantByType_Forecast_TotalLow = ""
    NDISParticipantByType_Forecast_TotalHigh = ""

class OBJ6:
    NDISTotalWorkforce_Forecast_Key = ""
    NDISTotalWorkforce_Forecast_Low = ""
    NDISTotalWorkforce_Forecast_High = ""
    NDISTotalWorkforce_Forecast_TotalLow = ""
    NDISTotalWorkforce_Forecast_TotalHigh = ""

listOBJ = []
listOBJ1 = []
listOBJ2 = []
listOBJ3 = []
listOBJ4 = []
listOBJ5 = []
listOBJ6 = []
# listCities = ['0810','6064','5606','5115','5068','3021','3053','3000','2600','2900','2800','2165','2000','2010']
listCities = ['2165','2000','2010']
listError = []

for i in listCities:
    URL = 'https://blcw.dss.gov.au/api/demand/GetPopulationByPostcode/'+str(i)

    res = requests.request("GET", URL)

    if res.status_code == 200:
        output = json.loads(res.text)

        obj1 = OBJ1()
        obj1.Postcode = output['postcode']
        obj1.Population = output['population']

        URL = 'https://blcw.dss.gov.au/api/demand/GetCurrentDemanByPostcode/' + str(i)

        res = requests.request("GET", URL)

        if res.status_code == 200:
            output = json.loads(res.text)

            list = output['estimatedDemandCurrent']['values']

            lowValTotal = output['estimatedDemandCurrent']['lowValTotal']
            highValTotal = output['estimatedDemandCurrent']['highValTotal']

            if(len(list) > 0):

                for j in range(0, len(list)):

                    obj2 = OBJ2()

                    obj2.NDISSpendByServiceArea_Current_TotalLow = lowValTotal
                    obj2.NDISSpendByServiceArea_Current_TotalHigh = highValTotal

                    obj2.NDISSpendByServiceArea_Current_Key = list[j]['key']
                    obj2.NDISSpendByServiceArea_Current_Low = list[j]['lowVal']
                    obj2.NDISSpendByServiceArea_Current_High = list[j]['highVal']


                    print(str(obj2.NDISSpendByServiceArea_Current_Key) + " | " + str(obj2.NDISSpendByServiceArea_Current_Low) + " | " + str(obj2.NDISSpendByServiceArea_Current_High))

                    listOBJ2.append(obj2)

            list = []

            list = output['participantsCurrent']['values']

            lowValTotal = output['participantsCurrent']['lowValTotal']
            highValTotal = output['participantsCurrent']['highValTotal']

            if(len(list) > 0):

                for j in range(0, len(list)):

                    obj2 = OBJ3()

                    obj2.NDISParticipantByType_Current_TotalLow = lowValTotal
                    obj2.NDISParticipantByType_Current_TotalHigh = highValTotal

                    obj2.NDISParticipantByType_Current_Key = list[j]['key']
                    obj2.NDISParticipantByType_Current_Low = list[j]['lowVal']
                    obj2.NDISParticipantByType_Current_High = list[j]['highVal']

                    print(str(obj2.NDISParticipantByType_Current_Key) + " | " + str(obj2.NDISParticipantByType_Current_Low) + " | " + str(obj2.NDISParticipantByType_Current_High))

                    listOBJ3.append(obj2)

        URL = 'https://blcw.dss.gov.au/api/demand/GetDemanByPostcode/' + str(i)

        res = requests.request("GET", URL)

        if res.status_code == 200:
            output = json.loads(res.text)

            list = output['estimatedDemand']['values']

            lowValTotal = output['estimatedDemand']['lowValTotal']
            highValTotal = output['estimatedDemand']['highValTotal']

            if(len(list) > 0):

                for j in range(0, len(list)):

                    obj2 = OBJ4()

                    obj2.NDISSpendByServiceArea_Forecast_TotalLow = lowValTotal
                    obj2.NDISSpendByServiceArea_Forecast_TotalHigh = highValTotal

                    obj2.NDISSpendByServiceArea_Forecast_Key = list[j]['key']
                    obj2.NDISSpendByServiceArea_Forecast_Low = list[j]['lowVal']
                    obj2.NDISSpendByServiceArea_Forecast_High = list[j]['highVal']

                    print(str(obj2.NDISSpendByServiceArea_Forecast_Key) + " | " + str(obj2.NDISSpendByServiceArea_Forecast_Low) + " | " + str(obj2.NDISSpendByServiceArea_Forecast_High))

                    listOBJ4.append(obj2)

            list = []

            list = output['participantProfile']['values']

            lowValTotal = output['participantProfile']['lowValTotal']
            highValTotal = output['participantProfile']['highValTotal']

            if(len(list) > 0):

                for j in range(0, len(list)):

                    obj2 = OBJ5()

                    obj2.NDISParticipantByType_Forecast_TotalLow = lowValTotal
                    obj2.NDISParticipantByType_Forecast_TotalHigh = highValTotal

                    obj2.NDISParticipantByType_Forecast_Key = list[j]['key']
                    obj2.NDISParticipantByType_Forecast_Low = list[j]['lowVal']
                    obj2.NDISParticipantByType_Forecast_High = list[j]['highVal']

                    print(str(obj2.NDISParticipantByType_Forecast_Key) + " | " + str(obj2.NDISParticipantByType_Forecast_Low) + " | " + str(obj2.NDISParticipantByType_Forecast_High))

                    listOBJ5.append(obj2)

            list = []

            list = output['totalWorkforce']['values']

            lowValTotal = output['totalWorkforce']['lowValTotal']
            highValTotal = output['totalWorkforce']['highValTotal']

            if(len(list) > 0):

                for j in range(0, len(list)):

                    obj2 = OBJ6()

                    obj2.NDISTotalWorkforce_Forecast_TotalLow = lowValTotal
                    obj2.NDISTotalWorkforce_Forecast_TotalHigh = highValTotal

                    obj2.NDISTotalWorkforce_Forecast_Key = list[j]['key']
                    obj2.NDISTotalWorkforce_Forecast_Low = list[j]['lowVal']
                    obj2.NDISTotalWorkforce_Forecast_High = list[j]['highVal']

                    print(str(obj2.NDISTotalWorkforce_Forecast_Key) + " | " + str(obj2.NDISTotalWorkforce_Forecast_Low) + " | " + str(obj2.NDISTotalWorkforce_Forecast_High))

                    listOBJ6.append(obj2)

        lisLength = [len(listOBJ2),len(listOBJ3),len(listOBJ4),len(listOBJ5),len(listOBJ6)]

        length = max(lisLength)

        for q in range(0, length):
            obj = OBJ()

            obj.Postcode = obj1.Postcode
            obj.Population = obj1.Population

            if len(listOBJ2) > q:
                obj.NDISSpendByServiceArea_Current_Key = listOBJ2[q].NDISSpendByServiceArea_Current_Key
                obj.NDISSpendByServiceArea_Current_Low = listOBJ2[q].NDISSpendByServiceArea_Current_Low
                obj.NDISSpendByServiceArea_Current_High = listOBJ2[q].NDISSpendByServiceArea_Current_High
                obj.NDISSpendByServiceArea_Current_TotalLow = listOBJ2[q].NDISSpendByServiceArea_Current_TotalLow
                obj.NDISSpendByServiceArea_Current_TotalHigh = listOBJ2[q].NDISSpendByServiceArea_Current_TotalHigh
            else:
                obj.NDISSpendByServiceArea_Current_Key = ""
                obj.NDISSpendByServiceArea_Current_Low = ""
                obj.NDISSpendByServiceArea_Current_High = ""
                obj.NDISSpendByServiceArea_Current_TotalLow = ""
                obj.NDISSpendByServiceArea_Current_TotalHigh = ""
            if len(listOBJ3) > q:
                obj.NDISParticipantByType_Current_Key = listOBJ3[q].NDISParticipantByType_Current_Key
                obj.NDISParticipantByType_Current_Low = listOBJ3[q].NDISParticipantByType_Current_Low
                obj.NDISParticipantByType_Current_High = listOBJ3[q].NDISParticipantByType_Current_High
                obj.NDISParticipantByType_Current_TotalLow = listOBJ3[q].NDISParticipantByType_Current_TotalLow
                obj.NDISParticipantByType_Current_TotalHigh = listOBJ3[q].NDISParticipantByType_Current_TotalHigh
            else:
                obj.NDISParticipantByType_Current_Key = ""
                obj.NDISParticipantByType_Current_Low = ""
                obj.NDISParticipantByType_Current_High = ""
                obj.NDISParticipantByType_Current_TotalLow = ""
                obj.NDISParticipantByType_Current_TotalHigh = ""
            if len(listOBJ4) > q:
                obj.NDISSpendByServiceArea_Forecast_Key = listOBJ4[q].NDISSpendByServiceArea_Forecast_Key
                obj.NDISSpendByServiceArea_Forecast_Low = listOBJ4[q].NDISSpendByServiceArea_Forecast_Low
                obj.NDISSpendByServiceArea_Forecast_High = listOBJ4[q].NDISSpendByServiceArea_Forecast_High
                obj.NDISSpendByServiceArea_Forecast_TotalLow = listOBJ4[q].NDISSpendByServiceArea_Forecast_TotalLow
                obj.NDISSpendByServiceArea_Forecast_TotalHigh = listOBJ4[q].NDISSpendByServiceArea_Forecast_TotalHigh
            else:
                obj.NDISSpendByServiceArea_Forecast_Key = ""
                obj.NDISSpendByServiceArea_Forecast_Low = ""
                obj.NDISSpendByServiceArea_Forecast_High = ""
                obj.NDISSpendByServiceArea_Forecast_TotalLow = ""
                obj.NDISSpendByServiceArea_Forecast_TotalHigh = ""
            if len(listOBJ5) > q:
                obj.NDISParticipantByType_Forecast_Key = listOBJ5[q].NDISParticipantByType_Forecast_Key
                obj.NDISParticipantByType_Forecast_Low = listOBJ5[q].NDISParticipantByType_Forecast_Low
                obj.NDISParticipantByType_Forecast_High = listOBJ5[q].NDISParticipantByType_Forecast_High
                obj.NDISParticipantByType_Forecast_TotalLow = listOBJ5[q].NDISParticipantByType_Forecast_TotalLow
                obj.NDISParticipantByType_Forecast_TotalHigh = listOBJ5[q].NDISParticipantByType_Forecast_TotalHigh
            else:
                obj.NDISParticipantByType_Forecast_Key = ""
                obj.NDISParticipantByType_Forecast_Low = ""
                obj.NDISParticipantByType_Forecast_High = ""
                obj.NDISParticipantByType_Forecast_TotalLow = ""
                obj.NDISParticipantByType_Forecast_TotalHigh = ""
            if len(listOBJ6) > q:
                obj.NDISTotalWorkforce_Forecast_Key = listOBJ6[q].NDISTotalWorkforce_Forecast_Key
                obj.NDISTotalWorkforce_Forecast_Low = listOBJ6[q].NDISTotalWorkforce_Forecast_Low
                obj.NDISTotalWorkforce_Forecast_High = listOBJ6[q].NDISTotalWorkforce_Forecast_High
                obj.NDISTotalWorkforce_Forecast_TotalLow = listOBJ6[q].NDISTotalWorkforce_Forecast_TotalLow
                obj.NDISTotalWorkforce_Forecast_TotalHigh = listOBJ6[q].NDISTotalWorkforce_Forecast_TotalHigh
            else:
                obj.NDISTotalWorkforce_Forecast_Key = ""
                obj.NDISTotalWorkforce_Forecast_Low = ""
                obj.NDISTotalWorkforce_Forecast_High = ""
                obj.NDISTotalWorkforce_Forecast_TotalLow = ""
                obj.NDISTotalWorkforce_Forecast_TotalHigh = ""

            print(str(q) + " of " + str(length) + " | "+str(obj.Postcode) + " | " + str(obj.Population))

            listOBJ.append(obj)



j = 1

sheet_obj_w.cell(row=j, column=1).value = 'Postcode'
sheet_obj_w.cell(row=j, column=2).value = 'Population'
sheet_obj_w.cell(row=j, column=3).value = 'NDISSpendByServiceArea_Current_Key'
sheet_obj_w.cell(row=j, column=4).value = 'NDISSpendByServiceArea_Current_Low'
sheet_obj_w.cell(row=j, column=5).value = 'NDISSpendByServiceArea_Current_High'
sheet_obj_w.cell(row=j, column=6).value = 'NDISSpendByServiceArea_Current_TotalLow'
sheet_obj_w.cell(row=j, column=7).value = 'NDISSpendByServiceArea_Current_TotalHigh'
sheet_obj_w.cell(row=j, column=8).value = 'NDISParticipantByType_Current_Key'
sheet_obj_w.cell(row=j, column=9).value = 'NDISParticipantByType_Current_Low'
sheet_obj_w.cell(row=j, column=10).value = 'NDISParticipantByType_Current_High'
sheet_obj_w.cell(row=j, column=11).value = 'NDISParticipantByType_Current_TotalLow'
sheet_obj_w.cell(row=j, column=12).value = 'NDISParticipantByType_Current_TotalHigh'
sheet_obj_w.cell(row=j, column=13).value = 'NDISSpendByServiceArea_Forecast_Key'
sheet_obj_w.cell(row=j, column=14).value = 'NDISSpendByServiceArea_Forecast_Low'
sheet_obj_w.cell(row=j, column=15).value = 'NDISSpendByServiceArea_Forecast_High'
sheet_obj_w.cell(row=j, column=16).value = 'NDISSpendByServiceArea_Forecast_TotalLow'
sheet_obj_w.cell(row=j, column=17).value = 'NDISSpendByServiceArea_Forecast_TotalHigh'
sheet_obj_w.cell(row=j, column=18).value = 'NDISParticipantByType_Forecast_Key'
sheet_obj_w.cell(row=j, column=19).value = 'NDISParticipantByType_Forecast_Low'
sheet_obj_w.cell(row=j, column=20).value = 'NDISParticipantByType_Forecast_High'
sheet_obj_w.cell(row=j, column=21).value = 'NDISParticipantByType_Forecast_TotalLow'
sheet_obj_w.cell(row=j, column=22).value = 'NDISParticipantByType_Forecast_TotalHigh'
sheet_obj_w.cell(row=j, column=23).value = 'NDISTotalWorkforce_Forecast_Key'
sheet_obj_w.cell(row=j, column=24).value = 'NDISTotalWorkforce_Forecast_Low'
sheet_obj_w.cell(row=j, column=25).value = 'NDISTotalWorkforce_Forecast_High'
sheet_obj_w.cell(row=j, column=26).value = 'NDISTotalWorkforce_Forecast_TotalLow'
sheet_obj_w.cell(row=j, column=27).value = 'NDISTotalWorkforce_Forecast_TotalHigh'
wb_obj_w.save("Documents\BLCW_AUS_18062021.xlsx")

for z in range(len(listOBJ)):
    j = j + 1
    print(str(j) + " of " +str(len(listOBJ)) +" | "+ str(listOBJ[z].Postcode) + " | " + str(listOBJ[z].Population))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Postcode)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Population)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].NDISSpendByServiceArea_Current_Key)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].NDISSpendByServiceArea_Current_Low)
    sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].NDISSpendByServiceArea_Current_High)
    sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].NDISSpendByServiceArea_Current_TotalLow)
    sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].NDISSpendByServiceArea_Current_TotalHigh)
    sheet_obj_w.cell(row=j, column=8).value = str(listOBJ[z].NDISParticipantByType_Current_Key)
    sheet_obj_w.cell(row=j, column=9).value = str(listOBJ[z].NDISParticipantByType_Current_Low)
    sheet_obj_w.cell(row=j, column=10).value = str(listOBJ[z].NDISParticipantByType_Current_High)
    sheet_obj_w.cell(row=j, column=11).value = str(listOBJ[z].NDISParticipantByType_Current_TotalLow)
    sheet_obj_w.cell(row=j, column=12).value = str(listOBJ[z].NDISParticipantByType_Current_TotalHigh)
    sheet_obj_w.cell(row=j, column=13).value = str(listOBJ[z].NDISSpendByServiceArea_Forecast_Key)
    sheet_obj_w.cell(row=j, column=14).value = str(listOBJ[z].NDISSpendByServiceArea_Forecast_Low)
    sheet_obj_w.cell(row=j, column=15).value = str(listOBJ[z].NDISSpendByServiceArea_Forecast_High)
    sheet_obj_w.cell(row=j, column=16).value = str(listOBJ[z].NDISSpendByServiceArea_Forecast_TotalLow)
    sheet_obj_w.cell(row=j, column=17).value = str(listOBJ[z].NDISSpendByServiceArea_Forecast_TotalHigh)
    sheet_obj_w.cell(row=j, column=18).value = str(listOBJ[z].NDISParticipantByType_Forecast_Key)
    sheet_obj_w.cell(row=j, column=19).value = str(listOBJ[z].NDISParticipantByType_Forecast_Low)
    sheet_obj_w.cell(row=j, column=20).value = str(listOBJ[z].NDISParticipantByType_Forecast_High)
    sheet_obj_w.cell(row=j, column=21).value = str(listOBJ[z].NDISParticipantByType_Forecast_TotalLow)
    sheet_obj_w.cell(row=j, column=22).value = str(listOBJ[z].NDISParticipantByType_Forecast_TotalHigh)
    sheet_obj_w.cell(row=j, column=23).value = str(listOBJ[z].NDISTotalWorkforce_Forecast_Key)
    sheet_obj_w.cell(row=j, column=24).value = str(listOBJ[z].NDISTotalWorkforce_Forecast_Low)
    sheet_obj_w.cell(row=j, column=25).value = str(listOBJ[z].NDISTotalWorkforce_Forecast_High)
    sheet_obj_w.cell(row=j, column=26).value = str(listOBJ[z].NDISTotalWorkforce_Forecast_TotalLow)
    sheet_obj_w.cell(row=j, column=27).value = str(listOBJ[z].NDISTotalWorkforce_Forecast_TotalHigh)
    wb_obj_w.save("Documents\BLCW_AUS_18062021.xlsx")


from bs4 import BeautifulSoup
import requests
import openpyxl
import requests 

import json
from types import SimpleNamespace

my_path = "Documents/Australian_All_Postcodes_11022021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row

print("Total Row: " + str(postcode_max_row))


my_path = "Documents/ALDI_AU_06052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

# for i in range(2, 3170):
#     print(str(int(float(postcode_sheet_obj.cell(row = i, column = 2).value))))

j = 0

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
listError = []
obj = OBJ()
listOBJ.append(obj)

for x in range(2, postcode_max_row+1):
    URL = 'https://storelocator.aldi.com.au/Presentation/AldiSued/en-AU/Search?LocX=' + str(postcode_sheet_obj.cell(row=x, column=3).value) + '&LocY=' + str(postcode_sheet_obj.cell(row=x, column=4).value)

    try:
        print(URL)

        res = requests.get(URL)

        if res.status_code == 200:
            soup = BeautifulSoup(res.content, "html.parser")

            try:
                ul = soup.find(id="search-results").find("ul", id="resultList")

                for li in ul:
                    try:
                        obj = OBJ()

                        obj.Title = li.find("strong", class_="resultItem-CompanyName").text
                        obj.Address = li.find("div", class_="resultItem-Street").text
                        obj.FullAddress = li.find("div", class_="resultItem-City").text


                        location = li['data-json']
                        location = json.loads(location)

                        obj.Latitude = str(location['locX'])
                        obj.Longitude = str(location['locY'])


                        obj.Country = "Australia"

                        print(obj.Title + " | " + obj.Address + " | " + obj.Country + " | " + str(
                            obj.Latitude) + " | " + str(obj.Longitude))

                        result = False

                        if len(listOBJ) > 0:
                            for i in range(len(listOBJ)):
                                if (str(obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(listOBJ[i].Longitude)):
                                    result = True
                                    break

                        if result == False:
                            listOBJ.append(obj)

                    except:
                        continue

            except:
                print("Exception: " + URL)
                listError.append(URL)
                continue

    except : 
        print("Exception: " + URL)
        listError.append(URL)
        continue

    if x > 5:
        break


for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" | "+listOBJ[z].Address +" | "+ str(listOBJ[z].Latitude) +" | "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row = j, column = 5).value = str(str(listOBJ[z].Latitude))
    sheet_obj_w.cell(row = j, column = 6).value = str(str(listOBJ[z].Longitude))
    wb_obj_w.save("Documents/ALDI_AU_06052021.xlsx")

j = j + 10

if (len(listError) > 0):
    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/ALDI_AU_06052021.xlsx")
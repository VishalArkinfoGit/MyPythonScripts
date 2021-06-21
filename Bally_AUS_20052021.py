from bs4 import BeautifulSoup
from math import trunc
from os import stat
import openpyxl
import requests
import json
from types import SimpleNamespace

# my_path = "Documents\Australian_All_Suburb_QLD_13052021.xlsx"
# postcode_wb_obj = openpyxl.load_workbook(my_path)
# postcode_sheet_obj = postcode_wb_obj.active
# postcode_max_col = postcode_sheet_obj.max_column
# postcode_max_row = postcode_sheet_obj.max_row
# print("Total Row: " + str(postcode_max_row))




my_path = "Documents\Bally_AUS_20052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active


class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""


listOBJ = []
listError = []

listLink = []

listPostcode = [2000,2047,2140,4218,3000,3148]

url = "https://www.bally.com.au/en_AU/store-locator?dwcont=C1526941302&dwfrm_storelocator_find=ok"
headers = {
  'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
  'cookie': 'dwanonymous_e1534bc8f716ecd955dd18162089fdbb=abiDlimgUSWI1AeLzJEicNuV2c; _gcl_au=1.1.162275779.1621074278; __cq_uuid=bceshCQa2WrAcApGcRGvKZO72S; __cq_seg=0~0.00!1~0.00!2~0.00!3~0.00!4~0.00!5~0.00!6~0.00!7~0.00!8~0.00!9~0.00; _ga=GA1.3.1469960817.1621074280; _qubitTracker=f2q2e4edzw8-0koplrsr3-z5ifwlk; qb_generic=:XlvjoOV:bally.com.au; _fbp=fb.2.1621074283949.367995612; _gid=GA1.3.1504435155.1621534202; firstVisitHeader=true; firstVisit=true; dwac_beea1f2a510784b1bfa978b168=pB73m7Sbo0minZzgSjEoK9lzbHfbO2hN520=|dw-only|||AUD|false|Australia/Sydney|true; cqcid=abiDlimgUSWI1AeLzJEicNuV2c; cquid=||; sid=pB73m7Sbo0minZzgSjEoK9lzbHfbO2hN520; shippingCountry=AU; __cq_dnt=0; dw_dnt=0; dwsid=N1rdVELLT35vx2rM7Hyyz0o3Rrgta4R6Kjcb3xcwHjx2HWRN0O1p1OjHmJZaRxcum4q1UAFgsgH1MUtrKpCUYQ==; AKA_A2=A; RT="z=1&dm=www.bally.com.au&si=cbec210f-1759-4a99-ad84-cf8187871411&ss=koyhgaqm&sl=0&tt=0"; dw=1; dw_cookies_accepted=1; dwresolutiondefined=true; qb_permanent=f2q2e4edzw8-0koplrsr3-z5ifwlk:44:4:6:8:0::0:1:0:Bgn6Fr:Bgp9Lx:A::::106.77.140.127:ahmedabad:148387:india:IN:23.03:72.6:surat metropolitan region:356008:gujarat:10002:migrated|1621074284899:::XmPj/24:XmPj/2N:0:0:0::0:0:bally.com.au:0; qb_session=4:1:12::0:XmPj/2N:0:0:0:0:bally.com.au',
  'dnt': '1',
  'origin': 'https://www.bally.com.au',
  'referer': 'https://www.bally.com.au/en_AU/store-locator?dwcont=C1526941302&dwfrm_storelocator_find=ok',
  'user-agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36',
  'x-requested-with': 'XMLHttpRequest'
}

# for i in range(1, postcode_max_row + 1):
#     # URL = 'https://www.bally.com.au/en_AU/store-locator#address='+str(postcode_sheet_obj.cell(row = i, column = 3).value)+'&format=ajax&latitude='+str(postcode_sheet_obj.cell(row = i, column = 4).value)+'&longitude='+str(postcode_sheet_obj.cell(row = i, column = 5).value)
#
#     payload = {'address': str(postcode_sheet_obj.cell(row = i, column = 3).value)+', Australia',
#                'format': 'ajax',
#                'country': 'AU',
#                 'addressValue': str(postcode_sheet_obj.cell(row = i, column = 3).value)}
#
#     print(str(i) + " of "+str(postcode_max_row)+" | " + str(postcode_sheet_obj.cell(row = i, column = 3).value) + " | " + str(postcode_sheet_obj.cell(row = i, column = 4).value) + " | " + str(postcode_sheet_obj.cell(row = i, column = 5).value) )

i=0
for code in listPostcode:
    # URL = 'https://www.bally.com.au/en_AU/store-locator#address='+str(postcode_sheet_obj.cell(row = i, column = 3).value)+'&format=ajax&latitude='+str(postcode_sheet_obj.cell(row = i, column = 4).value)+'&longitude='+str(postcode_sheet_obj.cell(row = i, column = 5).value)
    i = i+1
    payload = {'address': str(code)+', Australia',
               'format': 'ajax',
               'country': 'AU',
                'addressValue': str(code)}

    print(str(i) + " of "+str(len(listPostcode))+" | " + str(code) )

    try:

        res = requests.request("POST", url, headers=headers, data=payload)

        if res.status_code == 200:
            soup = BeautifulSoup(res.content, "html.parser")

            try:
                list = soup.find_all('div', class_='js-store-information')

                if len(list) > 0:
                    for div in list:
                        try:
                            zz = div['data-marker-info']

                            obj = OBJ()

                            obj.Title = str(div.find('div', class_='store-name').find('a').text).replace('\n','').replace('  ', '').strip()

                            output = json.loads(zz, object_hook=lambda d: SimpleNamespace(**d))

                            obj.Address = str(output.address)

                            try:
                                yz = obj.Address.split(',')
                                yz = yz[::-1]
                                obj.Address = ', '.join(yz)
                            except:
                                obj.Address = obj.Address

                            obj.Latitude = str(output.latitude)
                            obj.Longitude = str(output.longitude)

                            print(obj.Title + " | " + obj.Address + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                            result = False

                            if len(listOBJ) > 0:
                                for i in range(len(listOBJ)):
                                    if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(
                                            listOBJ[i].Address) and str(obj.Latitude) == str(
                                        listOBJ[i].Latitude) and str(
                                        obj.Longitude) == str(
                                        listOBJ[i].Longitude)):
                                        result = True
                                        break

                            if result == False:
                                listOBJ.append(obj)
                        except:
                            continue
                else:
                    continue

            except:
                continue
    except:
        continue

    # break

j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(str(j) + " of "+str(len(listOBJ))+" | " + listOBJ[z].Title + " | " + listOBJ[z].Address + " | " + str(listOBJ[z].Latitude) + " | " + str(
        listOBJ[z].Longitude))
    sheet_obj_w.cell(row=j, column=1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row=j, column=2).value = str(listOBJ[z].Address)
    # sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].FullAddress)
    # sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].City)
    # sheet_obj_w.cell(row=j, column=5).value = str(listOBJ[z].State)
    # sheet_obj_w.cell(row=j, column=6).value = str(listOBJ[z].Postcode)
    # sheet_obj_w.cell(row=j, column=7).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row=j, column=3).value = str(listOBJ[z].Latitude)
    sheet_obj_w.cell(row=j, column=4).value = str(listOBJ[z].Longitude)

    wb_obj_w.save("Documents\Bally_AUS_20052021.xlsx")

j = j + 10

if (len(listError) > 0):

    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents\Bally_AUS_20052021.xlsx")
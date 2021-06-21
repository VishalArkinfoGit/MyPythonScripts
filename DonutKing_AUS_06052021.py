from bs4 import BeautifulSoup
import requests
import openpyxl
import requests 

import json
from types import SimpleNamespace

my_path = "Documents/Australian_All_Postcodes_11022021.xlsx"
postcode_wb_obj = openpyxl.load_workbook(my_path)
postcode_sheet_obj = postcode_wb_obj.active
postcode_max_col = postcode_sheet_obj.max_column
postcode_max_row = postcode_sheet_obj.max_row

my_path = "Documents/DonutKing_AUS_06052021.xlsx"
wb_obj_w = openpyxl.load_workbook(my_path)
sheet_obj_w = wb_obj_w.active

class OBJ:
    Title = ""
    Address = ""
    FullAddress = ""
    Suburb = ""
    State = ""
    City = ""
    Country = ""
    Postcode = ""
    Latitude = ""
    Longitude = ""

listOBJ = []
listError = []

for x in range(2, postcode_max_row+1):
    URL = 'https://www.donutking.com.au/wp/wp-admin/admin-ajax.php?action=store_search&lat=' + str(postcode_sheet_obj.cell(row=x, column=3).value) + '&lng=' + str(postcode_sheet_obj.cell(row=x, column=4).value) + '&max_results=50&search_radius=500&autoload=1'
    try:
        print(URL)

        response = requests.get(URL).json()

        if response.status_code == 200:

            for y in range(len(response)):
                try:
                    x = response[y]

                    obj = OBJ()
                    obj.Title = str(x['store'])
                    obj.Address = str(x['address'])

                    try:
                        obj.FullAddress = str(x['address2'])
                    except:
                        obj.FullAddress = ''

                    obj.City = str(x['city'])
                    obj.State = str(x['state'])
                    obj.Postcode = str(x['zip'])
                    obj.Country = str(x['country'])
                    obj.Latitude = str(x['lat'])
                    obj.Longitude = str(x['lng'])

                    print(obj.Title + " | " + obj.Address + " | " + str(obj.Latitude) + " | " + str(obj.Longitude))

                    result = False

                    if len(listOBJ) > 0:
                        for i in range(len(listOBJ)):
                            if (str(obj.Title) == str(listOBJ[i].Title) and str(obj.Address) == str(listOBJ[i].Address) and str(obj.Latitude) == str(listOBJ[i].Latitude) and str(obj.Longitude) == str(listOBJ[i].Longitude)):
                                result = True
                                break

                    if result == False:
                        listOBJ.append(obj)


                except:
                    continue

    except : 
        print("Exception: " + URL)
        listError.append(URL)
        continue


j = 0

for z in range(len(listOBJ)):
    j = j + 1
    print(listOBJ[z].Title +" | "+listOBJ[z].Address +" | "+ str(listOBJ[z].Latitude) +" | "+ str(listOBJ[z].Longitude))
    sheet_obj_w.cell(row = j, column = 1).value = str(listOBJ[z].Title)
    sheet_obj_w.cell(row = j, column = 2).value = str(listOBJ[z].Address)
    sheet_obj_w.cell(row = j, column = 3).value = str(listOBJ[z].FullAddress)
    sheet_obj_w.cell(row = j, column = 4).value = str(listOBJ[z].Country)
    sheet_obj_w.cell(row = j, column = 5).value = str(str(listOBJ[z].Latitude))
    sheet_obj_w.cell(row = j, column = 6).value = str(str(listOBJ[z].Longitude))
    wb_obj_w.save("Documents/DonutKing_AUS_06052021.xlsx")

j = j + 10

if(len(listError) > 0):
    sheet_obj_w.cell(row=j, column=1).value = 'ERROR'

    for z in listError:
        j = j + 1
        sheet_obj_w.cell(row=j, column=1).value = str(z)
        wb_obj_w.save("Documents/DonutKing_AUS_06052021.xlsx")